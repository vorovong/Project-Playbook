"""미식가의 주방 정산 구글시트 v2
— 세로형 입력 폼(1개) + 전송 버튼 + 데이터 축적 + 정산 자동"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference

# ===== Styles =====
YELLOW = PatternFill("solid", fgColor="FFFDE7")
GRAY = PatternFill("solid", fgColor="F0F0F0")
BLUE_L = PatternFill("solid", fgColor="E3F2FD")
BLUE_D = PatternFill("solid", fgColor="1565C0")
GREEN_L = PatternFill("solid", fgColor="E8F5E9")
GREEN_BTN = PatternFill("solid", fgColor="4CAF50")
SECTION_BG = PatternFill("solid", fgColor="F5F5F5")
BOLD = Font(bold=True, size=11)
TITLE = Font(bold=True, size=14)
FORM_TITLE = Font(bold=True, size=15, color="1565C0")
SECTION_F = Font(bold=True, size=11, color="424242")
WHITE_B = Font(bold=True, size=11, color="FFFFFF")
HDR = Font(bold=True, size=10)
SMALL = Font(size=9, color="666666")
HINT = Font(size=9, color="999999", italic=True)
LABEL = Font(size=11, color="333333")
BTN_F = Font(bold=True, size=14, color="FFFFFF")
NUM = '#,##0'
BDR = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
BDR_INPUT = Border(
    left=Side('medium', color='1565C0'), right=Side('medium', color='1565C0'),
    top=Side('medium', color='1565C0'), bottom=Side('medium', color='1565C0'))
CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center')
RIGHT = Alignment(horizontal='right', vertical='center')

wb = openpyxl.Workbook()

# ============================================================
# 시트1: 주간 입력 (폼 1개 고정)
# ============================================================
ws1 = wb.active
ws1.title = "주간 입력"

ws1.column_dimensions['A'].width = 3
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 20
ws1.column_dimensions['D'].width = 22
ws1.column_dimensions['E'].width = 3

# 인쇄 영역 / 보기 설정
ws1.sheet_properties.tabColor = "1565C0"

# ---- 상단 타이틀 ----
ws1.merge_cells('B1:D1')
ws1['B1'] = '📋 미식가의 주방 — 주간 정산 입력'
ws1['B1'].font = FORM_TITLE
ws1['B1'].alignment = Alignment(horizontal='center')

ws1.merge_cells('B2:D2')
ws1['B2'] = '🟡 노란 칸에 입력 → 하단 [전송] 버튼 클릭'
ws1['B2'].font = SMALL
ws1['B2'].alignment = Alignment(horizontal='center')

r = 4  # 시작 행

# ---- 기본 정보 ----
ws1.merge_cells(f'B{r}:D{r}')
ws1.cell(r, 2, value='📅 기본 정보').font = SECTION_F
ws1.cell(r, 2).fill = SECTION_BG
ws1.cell(r, 3).fill = SECTION_BG
ws1.cell(r, 4).fill = SECTION_BG
r += 1

fields_basic = [
    ('월',      '1~12'),
    ('주차',    '1~5'),
    ('시작일',  '예: 3/3'),
    ('종료일',  '예: 3/9'),
]
input_cells = {}  # 입력 셀 위치 저장

for label, hint in fields_basic:
    ws1.cell(r, 2, value=label).font = LABEL
    ws1.cell(r, 2).alignment = RIGHT
    ws1.cell(r, 3).fill = YELLOW
    ws1.cell(r, 3).border = BDR_INPUT
    ws1.cell(r, 3).alignment = CTR
    if label in ['월', '주차']:
        ws1.cell(r, 3).font = Font(bold=True, size=13)
    ws1.cell(r, 4, value=f'← {hint}').font = HINT
    input_cells[label] = f'C{r}'
    r += 1

r += 1  # 빈줄

# ---- 매출 ----
ws1.merge_cells(f'B{r}:D{r}')
ws1.cell(r, 2, value='💰 매출').font = SECTION_F
ws1.cell(r, 2).fill = SECTION_BG
ws1.cell(r, 3).fill = SECTION_BG
ws1.cell(r, 4).fill = SECTION_BG
r += 1

fields_sales = [
    ('카드매출',  'POS 카드 매출액'),
    ('현금매출',  'POS 현금 매출액'),
]
for label, hint in fields_sales:
    ws1.cell(r, 2, value=label).font = LABEL
    ws1.cell(r, 2).alignment = RIGHT
    ws1.cell(r, 3).fill = YELLOW
    ws1.cell(r, 3).border = BDR_INPUT
    ws1.cell(r, 3).number_format = NUM
    ws1.cell(r, 3).font = Font(size=12)
    ws1.cell(r, 4, value=f'원  ({hint})').font = HINT
    input_cells[label] = f'C{r}'
    r += 1

r += 1

# ---- 공과금 ----
ws1.merge_cells(f'B{r}:D{r}')
ws1.cell(r, 2, value='⚡ 공과금  (3주차에 지난달분 입력)').font = SECTION_F
ws1.cell(r, 2).fill = SECTION_BG
ws1.cell(r, 3).fill = SECTION_BG
ws1.cell(r, 4).fill = SECTION_BG
r += 1

fields_utility = [
    ('전기요금',  '고지서 금액 그대로'),
    ('가스요금',  '고지서 금액 그대로'),
    ('수도요금',  '고지서 금액 그대로'),
]
for label, hint in fields_utility:
    ws1.cell(r, 2, value=label).font = LABEL
    ws1.cell(r, 2).alignment = RIGHT
    ws1.cell(r, 3).fill = YELLOW
    ws1.cell(r, 3).border = BDR_INPUT
    ws1.cell(r, 3).number_format = NUM
    ws1.cell(r, 4, value=f'원  ({hint})').font = HINT
    input_cells[label] = f'C{r}'
    r += 1

# 자동 계산 표시
ws1.cell(r, 2, value='→ 95% 자동계산').font = HINT
ws1.cell(r, 2).alignment = RIGHT
e = input_cells['전기요금']
g = input_cells['가스요금']
w = input_cells['수도요금']
ws1.cell(r, 3, value=f'=IF(AND({e}="",{g}="",{w}=""),"",(IF({e}="",0,{e})+IF({g}="",0,{g})+IF({w}="",0,{w}))*0.95)')
ws1.cell(r, 3).fill = GRAY
ws1.cell(r, 3).border = BDR
ws1.cell(r, 3).number_format = NUM
ws1.cell(r, 4, value='← 미식가 부담분').font = HINT
input_cells['공과금95'] = f'C{r}'
r += 1

r += 1

# ---- 기타 ----
ws1.merge_cells(f'B{r}:D{r}')
ws1.cell(r, 2, value='📝 기타  (발생할 때만)').font = SECTION_F
ws1.cell(r, 2).fill = SECTION_BG
ws1.cell(r, 3).fill = SECTION_BG
ws1.cell(r, 4).fill = SECTION_BG
r += 1

fields_etc = [
    ('정화조',    '연 2회 발생 시'),
    ('회식매출',  '재생전골 연계 등'),
    ('기타비용',  '그 외 비용'),
    ('비고',      '특이사항 메모'),
]
for label, hint in fields_etc:
    ws1.cell(r, 2, value=label).font = LABEL
    ws1.cell(r, 2).alignment = RIGHT
    ws1.cell(r, 3).fill = YELLOW
    ws1.cell(r, 3).border = BDR_INPUT
    if label != '비고':
        ws1.cell(r, 3).number_format = NUM
    ws1.cell(r, 4, value=hint).font = HINT
    input_cells[label] = f'C{r}'
    r += 1

r += 1

# ---- 전송 버튼 영역 ----
ws1.merge_cells(f'B{r}:D{r}')
ws1.cell(r, 2, value='▶  전송 (구글시트에서 버튼으로 전환)').font = BTN_F
ws1.cell(r, 2).fill = GREEN_BTN
ws1.cell(r, 2).alignment = CTR
ws1.cell(r, 3).fill = GREEN_BTN
ws1.cell(r, 4).fill = GREEN_BTN
for ci in [2,3,4]:
    ws1.cell(r, ci).border = Border(
        left=Side('medium'), right=Side('medium'),
        top=Side('medium'), bottom=Side('medium'))
btn_row = r
r += 1

ws1.merge_cells(f'B{r}:D{r}')
ws1.cell(r, 2, value='↑ 구글시트 업로드 후 Apps Script로 버튼 연결하세요').font = HINT
ws1.cell(r, 2).alignment = CTR
r += 2

# ---- 최근 입력 확인 (마지막 전송 데이터 미리보기) ----
ws1.merge_cells(f'B{r}:D{r}')
ws1.cell(r, 2, value='📌 최근 전송 기록 (최근 5건)').font = SECTION_F
ws1.cell(r, 2).fill = BLUE_L
ws1.cell(r, 3).fill = BLUE_L
ws1.cell(r, 4).fill = BLUE_L
r += 1

# 미니 테이블 헤더
for ci, h in enumerate(['월/주', '카드매출', '현금매출'], 2):
    ws1.cell(r, ci, value=h).font = HDR
    ws1.cell(r, ci).border = BDR
    ws1.cell(r, ci).alignment = CTR
    ws1.cell(r, ci).fill = GRAY
r += 1

# 최근 5건 참조 (데이터 시트에서)
for i in range(5):
    data_row = i + 2  # 데이터 시트 row 2부터
    ws1.cell(r, 2, value=f"=IF(데이터!A{data_row}=\"\",\"\",데이터!A{data_row}&\"월 \"&데이터!B{data_row}&\"주\")")
    ws1.cell(r, 2).border = BDR
    ws1.cell(r, 2).alignment = CTR
    ws1.cell(r, 3, value=f"=IF(데이터!A{data_row}=\"\",\"\",데이터!E{data_row})")
    ws1.cell(r, 3).border = BDR
    ws1.cell(r, 3).number_format = NUM
    ws1.cell(r, 4, value=f"=IF(데이터!A{data_row}=\"\",\"\",데이터!F{data_row})")
    ws1.cell(r, 4).border = BDR
    ws1.cell(r, 4).number_format = NUM
    r += 1

# 입력 셀 위치를 나중에 참조할 수 있도록 저장
# (Apps Script에서 사용)
INPUT_MAP = input_cells


# ============================================================
# 시트2: 데이터 (전송된 데이터 축적)
# ============================================================
ws_data = wb.create_sheet("데이터")
ws_data.sheet_properties.tabColor = "757575"

data_headers = [
    '월','주차','시작일','종료일',
    '카드매출','현금매출',
    '전기요금','가스요금','수도요금',
    '정화조','회식매출','기타비용','비고',
    '입력일시'
]
d_widths = [5,5,11,11, 15,15, 13,13,13, 13,13,13,18, 16]

for i, (h, w) in enumerate(zip(data_headers, d_widths), 1):
    ws_data.column_dimensions[get_column_letter(i)].width = w
    c = ws_data.cell(1, i, value=h)
    c.font = WHITE_B
    c.fill = PatternFill("solid", fgColor="424242")
    c.border = BDR
    c.alignment = CTR

# 빈 데이터 행 준비 (수식 없음 — Apps Script가 채움)
# row 2부터 데이터 쌓임


# ============================================================
# 시트3: 정산 요약 (데이터 시트 기반 자동 계산)
# ============================================================
ws3 = wb.create_sheet("정산 요약")
ws3.freeze_panes = 'A4'
ws3.sheet_properties.tabColor = "1565C0"

S3_COLS = 20
s3_widths = {
    'A':5,'B':5,'C':11,'D':11,
    'E':15,'F':15,'G':15,
    'H':14,'I':14,'J':14,'K':13,
    'L':14,'M':13,
    'N':13,'O':14,'P':13,
    'Q':16,'R':16,'S':14,'T':18
}
for c, w in s3_widths.items():
    ws3.column_dimensions[c].width = w

ws3.merge_cells('A1:T1')
ws3['A1'] = '미식가의 주방 — 주간 정산 요약 (2026년)'
ws3['A1'].font = TITLE
ws3['A1'].alignment = Alignment(horizontal='center')

ws3.merge_cells('A2:T2')
ws3['A2'] = '⬜ 전체 자동 — 입력 데이터에서 자동 계산   |   3주차=공과금 / 4주차=임대료+관리비'
ws3['A2'].font = SMALL
ws3['A2'].alignment = Alignment(horizontal='center')

s3_headers = [
    '월','주','시작일','종료일',
    '카드매출','현금매출','총매출',
    '카드수수료\n(15%)','현금수수료\n(13%)','수수료합계','수수료\nVAT',
    '임대료','임대료\nVAT',
    '관리비+\n고정비','공과금\n(95%)','기타비용',
    '총 차감액','미식가\n카드지급','현금\n수령액','비고'
]
for i, h in enumerate(s3_headers, 1):
    c = ws3.cell(3, i, value=h)
    c.font = HDR; c.border = BDR; c.alignment = CTR; c.fill = GRAY

# 데이터 시트에서 행을 끌어와 정산 계산
# 최대 60행 (12개월 × 5주)
row3 = 4
sub_rows_3 = []

for mi in range(12):
    mn = mi + 1
    fd3 = row3
    ld3 = row3 + 4

    for wk in range(1, 6):
        r3 = row3
        # 데이터 시트에서 해당 월/주차 데이터를 찾아오기
        # SUMPRODUCT + 조건으로 매칭 (월=mn, 주차=wk인 행)
        d_range_a = '데이터!$A$2:$A$200'   # 월
        d_range_b = '데이터!$B$2:$B$200'   # 주차
        cond = f'({d_range_a}={mn})*({d_range_b}={wk})'

        for ci in range(1, S3_COLS+1):
            ws3.cell(r3, ci).fill = GRAY
            ws3.cell(r3, ci).border = BDR

        # A, B: 고정값
        ws3.cell(r3, 1, value=mn).alignment = CTR
        ws3.cell(r3, 2, value=wk).alignment = CTR

        # C, D: 시작일/종료일
        ws3.cell(r3, 3, value=f'=IFERROR(SUMPRODUCT({cond}*데이터!$C$2:$C$200),"")')
        ws3.cell(r3, 4, value=f'=IFERROR(SUMPRODUCT({cond}*데이터!$D$2:$D$200),"")')

        # E: 카드매출
        ws3.cell(r3, 5, value=f'=SUMPRODUCT({cond}*데이터!$E$2:$E$200)')
        # F: 현금매출
        ws3.cell(r3, 6, value=f'=SUMPRODUCT({cond}*데이터!$F$2:$F$200)')
        # G: 총매출
        ws3.cell(r3, 7, value=f'=IF(E{r3}=0,"",E{r3}+F{r3})')

        # H~K: 수수료
        ws3.cell(r3, 8,  value=f'=IF(E{r3}=0,"",E{r3}*0.15)')
        ws3.cell(r3, 9,  value=f'=IF(E{r3}=0,"",F{r3}*0.13)')
        ws3.cell(r3, 10, value=f'=IF(E{r3}=0,"",H{r3}+I{r3})')
        ws3.cell(r3, 11, value=f'=IF(E{r3}=0,"",J{r3}*0.1)')

        # L~M: 임대료 (4주차에만)
        ws3.cell(r3, 12, value=f'=IF(AND(E{r3}<>0,B{r3}=4),3000000,0)')
        ws3.cell(r3, 13, value=f'=IF(L{r3}=0,0,L{r3}*0.1)')

        # N: 관리비+고정비 (4주차에만: 55000+15900+50000)
        ws3.cell(r3, 14, value=f'=IF(AND(E{r3}<>0,B{r3}=4),120900,0)')

        # O: 공과금 95%
        ws3.cell(r3, 15, value=f'=(SUMPRODUCT({cond}*데이터!$G$2:$G$200)+SUMPRODUCT({cond}*데이터!$H$2:$H$200)+SUMPRODUCT({cond}*데이터!$I$2:$I$200))*0.95')

        # P: 기타비용 (정화조 + 회식5% + 기타)
        ws3.cell(r3, 16, value=f'=SUMPRODUCT({cond}*데이터!$J$2:$J$200)+SUMPRODUCT({cond}*데이터!$K$2:$K$200)*0.05+SUMPRODUCT({cond}*데이터!$L$2:$L$200)')

        # Q: 총 차감액
        ws3.cell(r3, 17, value=f'=IF(E{r3}=0,"",J{r3}+K{r3}+L{r3}+M{r3}+N{r3}+O{r3}+P{r3})')

        # R: 미식가 카드지급
        ws3.cell(r3, 18, value=f'=IF(E{r3}=0,"",E{r3}-Q{r3})')

        # S: 현금 수령액
        ws3.cell(r3, 19, value=f'=IF(E{r3}=0,"",F{r3}-I{r3}-I{r3}*0.1)')

        # T: 비고
        ws3.cell(r3, 20, value=f'=IFERROR(INDEX(데이터!$M$2:$M$200,MATCH(1,{cond},0)),"")')

        for ci in [5,6,7,8,9,10,11,12,13,14,15,16,17,18,19]:
            ws3.cell(r3, ci).number_format = NUM

        row3 += 1

    # 월 소계
    sr3 = row3
    sub_rows_3.append(sr3)
    ws3.cell(sr3, 1, value=f'{mn}월').font = BOLD
    ws3.cell(sr3, 1).alignment = CTR
    ws3.cell(sr3, 2, value='소계').font = BOLD
    ws3.cell(sr3, 2).alignment = CTR
    for ci in range(1, S3_COLS+1):
        ws3.cell(sr3, ci).fill = BLUE_L
        ws3.cell(sr3, ci).font = BOLD
        ws3.cell(sr3, ci).border = BDR

    for ci in [5,6,7,8,9,10,11,12,13,14,15,16,17,18,19]:
        cl = get_column_letter(ci)
        ws3.cell(sr3, ci, value=f'=SUM({cl}{fd3}:{cl}{ld3})')
        ws3.cell(sr3, ci).number_format = NUM
    row3 += 1

# 연간 합계
tr3 = row3
ws3.cell(tr3, 1, value='연간').alignment = CTR
ws3.cell(tr3, 2, value='합계').alignment = CTR
for ci in range(1, S3_COLS+1):
    ws3.cell(tr3, ci).fill = PatternFill("solid", fgColor="1565C0")
    ws3.cell(tr3, ci).font = WHITE_B
    ws3.cell(tr3, ci).border = BDR
for ci in [5,6,7,8,9,10,11,12,13,14,15,16,17,18,19]:
    cl = get_column_letter(ci)
    refs = '+'.join([f'{cl}{r}' for r in sub_rows_3])
    ws3.cell(tr3, ci, value=f'={refs}')
    ws3.cell(tr3, ci).number_format = NUM


# ============================================================
# 시트4: 대시보드
# ============================================================
ws4 = wb.create_sheet("대시보드")
ws4.sheet_properties.tabColor = "4CAF50"
for c in 'ABCDEFGHIJKLMNO':
    ws4.column_dimensions[c].width = 13
ws4.column_dimensions['A'].width = 8

ws4.merge_cells('A1:M1')
ws4['A1'] = '미식가의 주방 — 경영 대시보드 (2026년)'
ws4['A1'].font = TITLE
ws4['A1'].alignment = Alignment(horizontal='center')

ws4.cell(2, 13, value='현재월 →').font = Font(bold=True, size=10, color="666666")
ws4.cell(2, 13).alignment = Alignment(horizontal='right')
ws4.cell(2, 14, value=3).fill = YELLOW
ws4.cell(2, 14).border = BDR
ws4.cell(2, 14).font = Font(bold=True, size=14)
ws4.cell(2, 14).alignment = CTR

kpi_data = [
    ('B', '이번달 총매출', '=INDEX(D9:D20,N2)'),
    ('E', '이번달 수수료수익', '=INDEX(G9:G20,N2)'),
    ('H', '이번달 임대수익', '=INDEX(H9:H20,N2)+INDEX(I9:I20,N2)'),
    ('K', '연간 누적매출', '=D21'),
]
for col_letter, label, formula in kpi_data:
    ci = openpyxl.utils.column_index_from_string(col_letter)
    ws4.merge_cells(f'{col_letter}3:{get_column_letter(ci+1)}3')
    c = ws4.cell(3, ci, value=label)
    c.font = Font(bold=True, size=9, color="666666")
    c.alignment = CTR
    ws4.merge_cells(f'{col_letter}4:{get_column_letter(ci+1)}4')
    c = ws4.cell(4, ci, value=formula)
    c.font = Font(bold=True, size=16, color="1565C0")
    c.alignment = CTR
    c.number_format = '#,##0"원"'
    for r in [3,4]:
        for cc in [ci, ci+1]:
            ws4.cell(r, cc).border = BDR

ws4.merge_cells('A7:M7')
ws4['A7'] = '📊 월별 정산 요약'
ws4['A7'].font = Font(bold=True, size=13)

sum_hdrs = ['월','카드매출','현금매출','총매출','카드수수료','현금수수료',
            '수수료합계','임대료','관리비','공과금','총차감','미식가지급','현금수령']
for i, h in enumerate(sum_hdrs, 1):
    c = ws4.cell(8, i, value=h)
    c.font = HDR; c.border = BDR; c.alignment = CTR; c.fill = GRAY

src_map = ['E','F','G','H','I','J','L','N','O','Q','R','S']

for mi in range(12):
    r = 9 + mi
    ws4.cell(r, 1, value=f'{mi+1}월').font = BOLD
    ws4.cell(r, 1).alignment = CTR
    ws4.cell(r, 1).border = BDR
    for ci, sc in enumerate(src_map, 2):
        c = ws4.cell(r, ci, value=f"='정산 요약'!{sc}{sub_rows_3[mi]}")
        c.number_format = NUM; c.border = BDR
        if ci in [11, 12]: c.fill = GREEN_L

ws4.cell(21, 1, value='합계').font = WHITE_B
ws4.cell(21, 1).alignment = CTR
for ci in range(1, 14):
    ws4.cell(21, ci).fill = PatternFill("solid", fgColor="1565C0")
    ws4.cell(21, ci).font = WHITE_B
    ws4.cell(21, ci).border = BDR
for ci in range(2, 14):
    cl = get_column_letter(ci)
    ws4.cell(21, ci, value=f'=SUM({cl}9:{cl}20)')
    ws4.cell(21, ci).number_format = NUM

# Charts
cats = Reference(ws4, min_col=1, min_row=9, max_row=20)

ch1 = BarChart()
ch1.type = "col"; ch1.grouping = "stacked"
ch1.title = "월별 매출 추이"; ch1.y_axis.title = "금액 (원)"
ch1.style = 10; ch1.width = 28; ch1.height = 14
ch1.add_data(Reference(ws4, min_col=2, min_row=8, max_row=20), titles_from_data=True)
ch1.add_data(Reference(ws4, min_col=3, min_row=8, max_row=20), titles_from_data=True)
ch1.set_categories(cats)
ch1.series[0].graphicalProperties.solidFill = "1565C0"
ch1.series[1].graphicalProperties.solidFill = "4CAF50"
ws4.add_chart(ch1, "A23")

ws4.cell(40, 1, value='수익 구성 데이터').font = SMALL
pie_items = [
    ('카드수수료', '=E21'), ('현금수수료', '=F21'),
    ('임대료', '=H21'), ('관리비', '=I21'), ('공과금', '=J21'),
]
for i, (label, formula) in enumerate(pie_items):
    ws4.cell(41+i, 1, value=label)
    ws4.cell(41+i, 2, value=formula)
    ws4.cell(41+i, 2).number_format = NUM

ch2 = PieChart()
ch2.title = "공존공간 수익 구성"; ch2.style = 10
ch2.width = 18; ch2.height = 14
ch2.add_data(Reference(ws4, min_col=2, min_row=40, max_row=45), titles_from_data=True)
ch2.set_categories(Reference(ws4, min_col=1, min_row=41, max_row=45))
ws4.add_chart(ch2, "H23")

ch3 = LineChart()
ch3.title = "월별 차감 vs 미식가 지급"; ch3.y_axis.title = "금액 (원)"
ch3.style = 10; ch3.width = 28; ch3.height = 14
ch3.add_data(Reference(ws4, min_col=11, min_row=8, max_row=20), titles_from_data=True)
ch3.add_data(Reference(ws4, min_col=12, min_row=8, max_row=20), titles_from_data=True)
ch3.set_categories(cats)
ch3.series[0].graphicalProperties.line.solidFill = "F44336"
ch3.series[1].graphicalProperties.line.solidFill = "4CAF50"
ws4.add_chart(ch3, "A39")


# ============================================================
# 시트5: 미수금 관리
# ============================================================
ws5 = wb.create_sheet("미수금 관리")
ws5.sheet_properties.tabColor = "F44336"
for c,w in {'A':5,'B':22,'C':16,'D':16,'E':16,'F':16,'G':10,'H':25}.items():
    ws5.column_dimensions[c].width = w

ws5.merge_cells('A1:H1')
ws5['A1'] = '미식가의 주방 — 미수금 관리'
ws5['A1'].font = TITLE
ws5['A1'].alignment = Alignment(horizontal='center')

ws5.merge_cells('A2:H2')
ws5['A2'] = '기존 미수금 현황 + 상환 계획'
ws5['A2'].font = SMALL
ws5['A2'].alignment = Alignment(horizontal='center')

for i, h in enumerate(['#','항목','원금','차감 누적','잔액','이자','상태','비고'], 1):
    c = ws5.cell(3, i, value=h)
    c.font = HDR; c.border = BDR; c.alignment = CTR; c.fill = GRAY

debts = [
    [1, '공과금 미수', 26480630, '', None, '', '미수', '2024.4~2025.12 누적'],
    [2, '임대료 미수', 6600000, '', None, '', '미수', ''],
    [3, '현금매출 수수료 미수', 2303154, '', None, '', '미수', ''],
    [4, '관리비 미수', 2123763, '', None, '', '미수', ''],
    [5, '정수기 임대료 미수', 333900, '', None, '', '미수', ''],
    [6, '이자 (연 10%)', 2766151, '', None, '', '미수', '지연이자율 연 10%'],
]

for idx, d in enumerate(debts):
    r = 4 + idx
    for ci, val in enumerate(d, 1):
        if val is None: continue
        c = ws5.cell(r, ci, value=val)
        c.border = BDR
        if ci in [3,4,5,6]: c.number_format = NUM
    ws5.cell(r, 4).fill = YELLOW
    ws5.cell(r, 5, value=f'=C{r}-IF(D{r}="",0,D{r})')
    ws5.cell(r, 5).number_format = NUM
    ws5.cell(r, 5).border = BDR

tr5 = 4 + len(debts)
for ci in range(1, 9):
    ws5.cell(tr5, ci).fill = BLUE_L; ws5.cell(tr5, ci).font = BOLD; ws5.cell(tr5, ci).border = BDR
ws5.cell(tr5, 2, value='합계')
for ci in [3,4,5,6]:
    cl = get_column_letter(ci)
    ws5.cell(tr5, ci, value=f'=SUM({cl}4:{cl}{tr5-1})')
    ws5.cell(tr5, ci).number_format = NUM

rp = tr5 + 2
ws5.merge_cells(f'A{rp}:H{rp}')
ws5.cell(rp, 1, value='📋 미수금 차감 계획').font = Font(bold=True, size=13)
for i, h in enumerate(['#','회차','예정일','차감금액','차감 후 잔액','','','비고'], 1):
    c = ws5.cell(rp+1, i, value=h)
    c.font = HDR; c.border = BDR; c.fill = GRAY

plans = [
    [1,'1회차','2026-03-02',6000000],[2,'2회차','2026-03-09',6000000],
    [3,'3회차','2026-03-16',6000000],[4,'4회차','2026-03-23',6000000],
    [5,'5회차','2026-03-30',6000000],[6,'6회차','2026-04-06',6000000],
    [7,'7회차','2026-04-20',4607598],
]
for idx, p in enumerate(plans):
    r = rp + 2 + idx
    for ci, val in enumerate(p, 1):
        c = ws5.cell(r, ci, value=val)
        c.border = BDR
        if ci == 4: c.number_format = NUM; c.fill = YELLOW
    if idx == 0:
        ws5.cell(r, 5, value=f'=E{tr5}-D{r}')
    else:
        ws5.cell(r, 5, value=f'=E{r-1}-D{r}')
    ws5.cell(r, 5).number_format = NUM; ws5.cell(r, 5).border = BDR

# ---- 시트 순서 정리 ----
# 주간 입력 → 정산 요약 → 대시보드 → 미수금 → (데이터는 마지막)
wb.move_sheet("데이터", offset=3)

# Save
out = r'C:\Users\Administrator\projects\Project_Playbook\mini-projects\박승현\미식가의주방-정산\미식가의주방_정산_v2.xlsx'
wb.save(out)
print('Done!')
print(f'\n입력 셀 위치 (Apps Script용):')
for k, v in INPUT_MAP.items():
    print(f'  {k}: {v}')
