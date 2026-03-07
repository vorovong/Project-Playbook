"""미식가의 주방 정산 구글시트 생성"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference

# ===== Styles =====
YELLOW = PatternFill("solid", fgColor="FFFDE7")
GRAY = PatternFill("solid", fgColor="F0F0F0")
BLUE_L = PatternFill("solid", fgColor="E3F2FD")
BLUE_D = PatternFill("solid", fgColor="1565C0")
GREEN_L = PatternFill("solid", fgColor="E8F5E9")
RED_L = PatternFill("solid", fgColor="FFEBEE")
BOLD = Font(bold=True, size=11)
TITLE = Font(bold=True, size=14)
WHITE_B = Font(bold=True, size=11, color="FFFFFF")
HDR = Font(bold=True, size=10)
SMALL = Font(size=9, color="666666")
NUM = '#,##0'
BDR = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_range(ws, row, cols, fill, font=None):
    for c in range(1, cols+1):
        cell = ws.cell(row, c)
        cell.fill = fill
        cell.border = BDR
        if font: cell.font = font

wb = openpyxl.Workbook()

# ============================================================
# 시트1: 정산 입력
# ============================================================
ws = wb.active
ws.title = "정산 입력"
ws.freeze_panes = 'A4'

# Column widths
for c, w in {'A':5,'B':5,'C':11,'D':11,'E':15,'F':15,'G':15,'H':14,'I':14,'J':14,'K':13,'L':14,'M':13,'N':12,'O':14,'P':12,'Q':12,'R':16,'S':16,'T':14,'U':20}.items():
    ws.column_dimensions[c].width = w

# Title
ws.merge_cells('A1:U1')
ws['A1'] = '미식가의 주방 — 주간 정산표 (2026년)'
ws['A1'].font = TITLE
ws['A1'].alignment = Alignment(horizontal='center')

ws.merge_cells('A2:U2')
ws['A2'] = '🟡 노란색 = 직접 입력   |   회색 = 자동 계산 (수정 금지)   |   🔵 파란색 = 월 소계'
ws['A2'].font = SMALL
ws['A2'].alignment = Alignment(horizontal='center')

# Headers
headers = ['월','주','시작일','종료일','카드매출','현금매출','총매출',
           '카드수수료\n(15%)','현금수수료\n(13%)','수수료합계','수수료\nVAT',
           '임대료\n(주간배분)','임대료\nVAT','관리비\n(주간배분)',
           '공과금','캡스/넷','기타비용',
           '총 차감액','미식가\n카드지급','현금\n수령액','비고']
INPUT_COLS = {1,2,3,4,5,6,15,17,21}

for i, h in enumerate(headers, 1):
    c = ws.cell(3, i, value=h)
    c.font = HDR
    c.border = BDR
    c.alignment = CTR
    c.fill = YELLOW if i in INPUT_COLS else GRAY

# Data rows
row = 4
sub_rows = []

for mi in range(12):
    mn = mi + 1
    fd = row          # first data row
    ld = row + 4      # last data row

    for wk in range(1, 6):
        r = row
        # Input cells
        for ci in INPUT_COLS:
            ws.cell(r, ci).fill = YELLOW
            ws.cell(r, ci).border = BDR
        ws.cell(r, 1, value=mn).alignment = CTR
        ws.cell(r, 2, value=wk).alignment = CTR
        for ci in [5,6,15,17]:
            ws.cell(r, ci).number_format = NUM

        # Auto-calc cells
        for ci in [7,8,9,10,11,12,13,14,16,18,19,20]:
            ws.cell(r, ci).fill = GRAY
            ws.cell(r, ci).border = BDR
            ws.cell(r, ci).number_format = NUM

        # Formulas
        ws.cell(r, 7,  value=f'=IF(E{r}="","",E{r}+F{r})')
        ws.cell(r, 8,  value=f'=IF(E{r}="","",E{r}*0.15)')
        ws.cell(r, 9,  value=f'=IF(E{r}="","",IF(F{r}="",0,F{r}*0.13))')
        ws.cell(r, 10, value=f'=IF(E{r}="","",H{r}+I{r})')
        ws.cell(r, 11, value=f'=IF(E{r}="","",J{r}*0.1)')
        ws.cell(r, 12, value=f'=IF(E{r}="","",3000000/COUNTA(E{fd}:E{ld}))')
        ws.cell(r, 13, value=f'=IF(E{r}="","",L{r}*0.1)')
        ws.cell(r, 14, value=f'=IF(E{r}="","",70900/COUNTA(E{fd}:E{ld}))')
        ws.cell(r, 16, value=f'=IF(E{r}="","",IF(B{r}=1,50000,0))')
        ws.cell(r, 18, value=f'=IF(E{r}="","",H{r}+I{r}+K{r}+L{r}+M{r}+N{r}+IF(O{r}="",0,O{r})+P{r}+IF(Q{r}="",0,Q{r}))')
        ws.cell(r, 19, value=f'=IF(E{r}="","",E{r}-(R{r}-I{r}*1.1))')
        ws.cell(r, 20, value=f'=IF(E{r}="","",I{r}*1.1)')
        row += 1

    # Monthly subtotal
    sr = row
    sub_rows.append(sr)
    ws.cell(sr, 1, value=f'{mn}월').font = BOLD
    ws.cell(sr, 1).alignment = CTR
    ws.cell(sr, 2, value='소계').font = BOLD
    ws.cell(sr, 2).alignment = CTR
    style_range(ws, sr, 21, BLUE_L, BOLD)

    for ci in [5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20]:
        cl = get_column_letter(ci)
        ws.cell(sr, ci, value=f'=SUM({cl}{fd}:{cl}{ld})')
        ws.cell(sr, ci).number_format = NUM
    row += 1

# Grand total
tr = row
ws.cell(tr, 1, value='연간').alignment = CTR
ws.cell(tr, 2, value='합계').alignment = CTR
style_range(ws, tr, 21, BLUE_D, WHITE_B)

for ci in [5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20]:
    cl = get_column_letter(ci)
    refs = '+'.join([f'{cl}{r}' for r in sub_rows])
    ws.cell(tr, ci, value=f'={refs}')
    ws.cell(tr, ci).number_format = NUM

# ============================================================
# 시트2: 대시보드
# ============================================================
ws2 = wb.create_sheet("대시보드")
for c in 'ABCDEFGHIJKLMNO':
    ws2.column_dimensions[c].width = 13
ws2.column_dimensions['A'].width = 8

# Title
ws2.merge_cells('A1:M1')
ws2['A1'] = '미식가의 주방 — 경영 대시보드 (2026년)'
ws2['A1'].font = TITLE
ws2['A1'].alignment = Alignment(horizontal='center')

# 현재월 입력
ws2.cell(2, 13, value='현재월 →').font = Font(bold=True, size=10, color="666666")
ws2.cell(2, 13).alignment = Alignment(horizontal='right')
ws2.cell(2, 14, value=3).fill = YELLOW
ws2.cell(2, 14).border = BDR
ws2.cell(2, 14).font = Font(bold=True, size=14)
ws2.cell(2, 14).alignment = CTR
ws2.column_dimensions['N'].width = 6

# KPI Cards
kpi_data = [
    ('B', '이번달 총매출', '=INDEX(D9:D20,N2)'),
    ('E', '이번달 수수료수익', '=INDEX(G9:G20,N2)'),
    ('H', '이번달 임대수익', '=INDEX(H9:H20,N2)+INDEX(I9:I20,N2)'),
    ('K', '연간 누적매출', f'=D21'),
]
for col_letter, label, formula in kpi_data:
    ci = openpyxl.utils.column_index_from_string(col_letter)
    ws2.merge_cells(f'{col_letter}3:{get_column_letter(ci+1)}3')
    c = ws2.cell(3, ci, value=label)
    c.font = Font(bold=True, size=9, color="666666")
    c.alignment = CTR

    ws2.merge_cells(f'{col_letter}4:{get_column_letter(ci+1)}4')
    c = ws2.cell(4, ci, value=formula)
    c.font = Font(bold=True, size=16, color="1565C0")
    c.alignment = CTR
    c.number_format = '#,##0"원"'

    for r in [3,4]:
        for cc in [ci, ci+1]:
            ws2.cell(r, cc).border = BDR

# Separator
ws2.merge_cells('A6:M6')

# Monthly summary table
ws2.merge_cells('A7:M7')
ws2['A7'] = '📊 월별 정산 요약'
ws2['A7'].font = Font(bold=True, size=13)

sum_hdrs = ['월','카드매출','현금매출','총매출','카드수수료','현금수수료',
            '수수료합계','임대료','관리비','공과금','총차감','미식가지급','현금수령']
for i, h in enumerate(sum_hdrs, 1):
    c = ws2.cell(8, i, value=h)
    c.font = HDR
    c.border = BDR
    c.alignment = CTR
    c.fill = GRAY

# Source columns in 정산입력 for each summary column
src_map = ['E','F','G','H','I','J','L','N','O','R','S','T']

for mi in range(12):
    r = 9 + mi
    ws2.cell(r, 1, value=f'{mi+1}월').font = BOLD
    ws2.cell(r, 1).alignment = CTR
    ws2.cell(r, 1).border = BDR

    for ci, sc in enumerate(src_map, 2):
        c = ws2.cell(r, ci, value=f"='정산 입력'!{sc}{sub_rows[mi]}")
        c.number_format = NUM
        c.border = BDR
        # Highlight negative values
        if ci in [11, 12]:  # 미식가지급
            c.fill = GREEN_L

# Annual total
ws2.cell(21, 1, value='합계').font = WHITE_B
ws2.cell(21, 1).alignment = CTR
style_range(ws2, 21, 13, BLUE_D, WHITE_B)
for ci in range(2, 14):
    cl = get_column_letter(ci)
    ws2.cell(21, ci, value=f'=SUM({cl}9:{cl}20)')
    ws2.cell(21, ci).number_format = NUM

# ---- Charts ----

# Chart 1: 월별 매출 추이 (Stacked Bar)
ch1 = BarChart()
ch1.type = "col"
ch1.grouping = "stacked"
ch1.title = "월별 매출 추이"
ch1.y_axis.title = "금액 (원)"
ch1.style = 10
ch1.width = 28
ch1.height = 14

cats = Reference(ws2, min_col=1, min_row=9, max_row=20)
ch1.add_data(Reference(ws2, min_col=2, min_row=8, max_row=20), titles_from_data=True)
ch1.add_data(Reference(ws2, min_col=3, min_row=8, max_row=20), titles_from_data=True)
ch1.set_categories(cats)
ch1.series[0].graphicalProperties.solidFill = "1565C0"
ch1.series[1].graphicalProperties.solidFill = "4CAF50"
ws2.add_chart(ch1, "A23")

# Chart 2: 수익 구성 (Pie)
# Helper data
ws2.cell(40, 1, value='수익 구성 데이터').font = SMALL
pie_items = [
    ('카드수수료', f'=E21'),
    ('현금수수료', f'=F21'),
    ('임대료', f'=H21'),
    ('관리비', f'=I21'),
    ('공과금', f'=J21'),
]
for i, (label, formula) in enumerate(pie_items):
    ws2.cell(41+i, 1, value=label)
    ws2.cell(41+i, 2, value=formula)
    ws2.cell(41+i, 2).number_format = NUM

ch2 = PieChart()
ch2.title = "공존공간 수익 구성"
ch2.style = 10
ch2.width = 18
ch2.height = 14
ch2.add_data(Reference(ws2, min_col=2, min_row=40, max_row=45), titles_from_data=True)
ch2.set_categories(Reference(ws2, min_col=1, min_row=41, max_row=45))
ws2.add_chart(ch2, "H23")

# Chart 3: 총차감 vs 지급 추이 (Line)
ch3 = LineChart()
ch3.title = "월별 차감 vs 미식가 지급"
ch3.y_axis.title = "금액 (원)"
ch3.style = 10
ch3.width = 28
ch3.height = 14

ch3.add_data(Reference(ws2, min_col=11, min_row=8, max_row=20), titles_from_data=True)
ch3.add_data(Reference(ws2, min_col=12, min_row=8, max_row=20), titles_from_data=True)
ch3.set_categories(cats)
ch3.series[0].graphicalProperties.line.solidFill = "F44336"
ch3.series[1].graphicalProperties.line.solidFill = "4CAF50"
ws2.add_chart(ch3, "A39")

# ============================================================
# 시트3: 미수금 관리
# ============================================================
ws3 = wb.create_sheet("미수금 관리")
for c,w in {'A':5,'B':22,'C':16,'D':16,'E':16,'F':16,'G':10,'H':25}.items():
    ws3.column_dimensions[c].width = w

ws3.merge_cells('A1:H1')
ws3['A1'] = '미식가의 주방 — 미수금 관리'
ws3['A1'].font = TITLE
ws3['A1'].alignment = Alignment(horizontal='center')

ws3.merge_cells('A2:H2')
ws3['A2'] = '기존 미수금 현황 + 상환 계획 (세금계산서 vs 은행거래 비교 기준)'
ws3['A2'].font = SMALL
ws3['A2'].alignment = Alignment(horizontal='center')

# Headers
for i, h in enumerate(['#','항목','원금','차감 누적','잔액','이자','상태','비고'], 1):
    c = ws3.cell(3, i, value=h)
    c.font = HDR; c.border = BDR; c.alignment = CTR; c.fill = GRAY

# Debt data
debts = [
    [1, '공과금 미수', 26480630, '', None, '', '미수', '2024.4~2025.12 누적 (가장 큰 항목)'],
    [2, '임대료 미수', 6600000, '', None, '', '미수', ''],
    [3, '현금매출 수수료 미수', 2303154, '', None, '', '미수', ''],
    [4, '관리비 미수', 2123763, '', None, '', '미수', ''],
    [5, '정수기 임대료 미수', 333900, '', None, '', '미수', ''],
    [6, '이자 (연 10%)', 2766151, '', None, '', '미수', '지연이자율 연 10%'],
]

for idx, d in enumerate(debts):
    r = 4 + idx
    for ci, val in enumerate(d, 1):
        if val is None: continue
        c = ws3.cell(r, ci, value=val)
        c.border = BDR
        if ci in [3,4,5,6]: c.number_format = NUM
    ws3.cell(r, 4).fill = YELLOW  # 차감 누적은 입력
    ws3.cell(r, 5, value=f'=C{r}-IF(D{r}="",0,D{r})')
    ws3.cell(r, 5).number_format = NUM
    ws3.cell(r, 5).border = BDR

# Total
tr3 = 4 + len(debts)
style_range(ws3, tr3, 8, BLUE_L, BOLD)
ws3.cell(tr3, 2, value='합계').font = BOLD
for ci in [3,4,5,6]:
    cl = get_column_letter(ci)
    ws3.cell(tr3, ci, value=f'=SUM({cl}4:{cl}{tr3-1})')
    ws3.cell(tr3, ci).number_format = NUM

# Repayment plan
rp = tr3 + 2
ws3.merge_cells(f'A{rp}:H{rp}')
ws3.cell(rp, 1, value='📋 미수금 차감 계획 (주간 정산에서 분할 차감)').font = Font(bold=True, size=13)

for i, h in enumerate(['#','회차','예정일','차감금액','차감 후 잔액','','','비고'], 1):
    c = ws3.cell(rp+1, i, value=h)
    c.font = HDR; c.border = BDR; c.fill = GRAY

plans = [
    [1, '1회차', '2026-03-02', 6000000],
    [2, '2회차', '2026-03-09', 6000000],
    [3, '3회차', '2026-03-16', 6000000],
    [4, '4회차', '2026-03-23', 6000000],
    [5, '5회차', '2026-03-30', 6000000],
    [6, '6회차', '2026-04-06', 6000000],
    [7, '7회차', '2026-04-20', 4607598],
]

total_debt_cell = f'E{tr3}'
for idx, p in enumerate(plans):
    r = rp + 2 + idx
    for ci, val in enumerate(p, 1):
        c = ws3.cell(r, ci, value=val)
        c.border = BDR
        if ci == 4:
            c.number_format = NUM
            c.fill = YELLOW
    # 잔액 계산
    if idx == 0:
        ws3.cell(r, 5, value=f'={total_debt_cell}-D{r}')
    else:
        ws3.cell(r, 5, value=f'=E{r-1}-D{r}')
    ws3.cell(r, 5).number_format = NUM
    ws3.cell(r, 5).border = BDR

# Save
out = r'C:\Users\Administrator\projects\Project_Playbook\mini-projects\박승현\미식가의주방-정산\미식가의주방_정산_구글시트.xlsx'
wb.save(out)
print('Done: ' + out)
