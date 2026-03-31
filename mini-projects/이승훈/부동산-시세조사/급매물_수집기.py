"""
소액 꼬마빌딩 급매물 수집기
- 서울/경기/인천 10억 이하 건물 매물 자동 수집
- 네이버 부동산 데이터 기반
- 매일 실행하여 신규 매물 탐지

사용법:
  python 급매물_수집기.py              # 전체 수집 + 엑셀 저장
  python 급매물_수집기.py --region 서울  # 특정 지역만
"""

import os
import sys
import json
import time
import random
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ── 설정 ──────────────────────────────────────────────

DELAY = 5.0  # API 호출 간 기본 딜레이 (초) — 차단 방지용
MAX_RETRIES = 2  # 실패 시 재시도 횟수 (넘으면 다음 지역으로)

# 텔레그램 알림 설정
TELEGRAM_TOKEN = "8620151069:AAEvOeFJuhNQG74yq2oRC_tew2clD57Fy2g"
TELEGRAM_CHAT_ID = "8375504457"


def create_session():
    """네이버 부동산 API용 세션 생성"""
    s = requests.Session()
    s.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ),
        "Referer": "https://m.land.naver.com/",
    })
    return s


SESSION = create_session()


def api_get(url, params, retry=0):
    """API 호출 (재시도 + 랜덤 딜레이 포함)"""
    try:
        resp = SESSION.get(url, params=params, timeout=15, allow_redirects=False)
        if resp.status_code in (302, 307):
            location = resp.headers.get("Location", "")
            if "abuse" in location or "error" in location:
                if retry < MAX_RETRIES:
                    wait_time = (retry + 1) * 30 + random.uniform(0, 10)
                    print(f"  (차단 감지, {wait_time:.0f}초 대기 후 재시도...)")
                    time.sleep(wait_time)
                    return api_get(url, params, retry + 1)
                return {}  # 차단 풀리지 않으면 빈 결과로 다음 지역 진행
        resp.raise_for_status()
        return resp.json()
    except requests.exceptions.JSONDecodeError:
        if retry < MAX_RETRIES:
            time.sleep(3)
            return api_get(url, params, retry + 1)
        return {}


def wait():
    """랜덤 딜레이"""
    time.sleep(DELAY + random.uniform(0, 1))


# ── 텔레그램 알림 ────────────────────────────────────

def send_telegram(text):
    """텔레그램 메시지 발송"""
    try:
        requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
            json={
                "chat_id": TELEGRAM_CHAT_ID,
                "text": text,
                "parse_mode": "HTML",
                "disable_web_page_preview": True,
            },
            timeout=10,
        )
    except Exception:
        pass


def send_telegram_file(filepath, caption=""):
    """텔레그램으로 파일 전송"""
    try:
        with open(filepath, "rb") as f:
            requests.post(
                f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendDocument",
                data={"chat_id": TELEGRAM_CHAT_ID, "caption": caption},
                files={"document": f},
                timeout=30,
            )
    except Exception:
        pass


def notify_buildings(buildings, excel_path=None):
    """텔레그램 알림 — 엑셀 파일만 첨부"""
    if not buildings:
        send_telegram("오늘 신규 꼬마빌딩 매물이 없습니다.")
        return

    if excel_path and os.path.exists(excel_path):
        send_telegram_file(excel_path)
    else:
        send_telegram(f"꼬마빌딩 {len(buildings)}건 수집 완료")

# 서울/경기/인천 시도 코드
TARGET_SIDO = {
    "서울": "1100000000",
    "경기": "4100000000",
    "인천": "2800000000",
}

# 꼬마빌딩 관련 매물 유형 코드 (네이버 부동산 rletTpCd)
BUILDING_SEARCH_CODES = ["DDDG", "SG"]

# 매물 상세에서 필터링할 유형명 (rletTpNm) — 꼬마빌딩(건물 통째)만
BUILDING_TYPE_NAMES = {"건물", "상가건물", "상가주택"}

# 가격 상한 (만원 단위, 10억 = 100000만원)
MAX_PRICE = 100000

# 저장 경로
DB_DIR = os.path.join(os.path.expanduser("~"), "OneDrive", "바탕 화면", "DB")
DATA_DIR = os.path.join(
    os.path.expanduser("~"),
    "projects", "Project_Playbook",
    "mini-projects", "이승훈", "부동산-시세조사", "data",
)


# ── 1. 지역 코드 수집 ────────────────────────────────

def get_regions(cortar_no):
    """cortarNo로 하위 지역 목록 조회"""
    data = api_get(
        "https://m.land.naver.com/map/getRegionList",
        params={"cortarNo": cortar_no},
    )
    return data.get("result", {}).get("list", [])


def get_all_gungu(sido_codes=None, log=print):
    """서울/경기/인천의 모든 시군구 목록 조회"""
    if sido_codes is None:
        sido_codes = TARGET_SIDO

    all_gungu = []
    for name, code in sido_codes.items():
        log(f"[지역] {name} 시군구 목록 조회 중...")
        gungu_list = get_regions(code)
        for g in gungu_list:
            all_gungu.append({
                "sido": name,
                "gungu": g["CortarNm"],
                "cortarNo": g["CortarNo"],
                "lat": float(g["MapYCrdn"]),
                "lon": float(g["MapXCrdn"]),
            })
        wait()
        log(f"  → {name}: {len(gungu_list)}개 시군구")

    return all_gungu


# ── 2. 매물 수집 ─────────────────────────────────────

def fetch_clusters(lat, lon, cortar_no, rlet_type, trade_type="A1"):
    """지역 좌표 기준으로 매물 클러스터 조회"""
    offset = 0.03
    data = api_get(
        "https://m.land.naver.com/cluster/clusterList",
        params={
            "view": "atcl",
            "rletTpCd": rlet_type,
            "tradTpCd": trade_type,
            "z": "14",
            "lat": str(lat),
            "lon": str(lon),
            "btm": str(lat - offset),
            "lft": str(lon - offset),
            "top": str(lat + offset),
            "rgt": str(lon + offset),
            "cortarNo": cortar_no,
        },
    )
    return data.get("data", {}).get("ARTICLE", [])


def fetch_articles_from_cluster(cluster, cortar_no, rlet_type, trade_type="A1"):
    """클러스터에서 개별 매물 목록 조회 (페이징 포함)"""
    all_articles = []
    page = 1
    offset = 0.03

    while True:
        data = api_get(
            "https://m.land.naver.com/cluster/ajax/articleList",
            params={
                "rletTpCd": rlet_type,
                "tradTpCd": trade_type,
                "z": "14",
                "lat": str(cluster["lat"]),
                "lon": str(cluster["lon"]),
                "btm": str(cluster["lat"] - offset),
                "lft": str(cluster["lon"] - offset),
                "top": str(cluster["lat"] + offset),
                "rgt": str(cluster["lon"] + offset),
                "lgeo": cluster["lgeo"],
                "showR0": "N",
                "page": str(page),
                "cortarNo": cortar_no,
            },
        )
        body = data.get("body", [])
        if not body:
            break
        all_articles.extend(body)
        if not data.get("more", False):
            break
        page += 1
        wait()

    return all_articles


def is_building_type(article):
    """꼬마빌딩에 해당하는 매물인지 판별"""
    tp = article.get("rletTpNm", "")
    return tp in BUILDING_TYPE_NAMES


def is_recent(article, days=1):
    """최근 N일 이내 등록된 매물인지 판별 (cfmYmd 기준)"""
    cfm = article.get("cfmYmd", "")
    if not cfm:
        return True  # 날짜 없으면 일단 포함
    try:
        from datetime import timedelta
        cfm_date = datetime.strptime(cfm, "%y.%m.%d").date()
        cutoff = datetime.now().date() - timedelta(days=days)
        return cfm_date >= cutoff
    except (ValueError, TypeError):
        return True


def parse_article(article):
    """매물 데이터를 정리된 딕셔너리로 변환"""
    prc = int(article.get("prc", 0) or 0)

    try:
        spc1 = float(article.get("spc1", 0) or 0)
    except (ValueError, TypeError):
        spc1 = 0
    try:
        spc2 = float(article.get("spc2", 0) or 0)
    except (ValueError, TypeError):
        spc2 = 0

    area = spc2 if spc2 > 0 else spc1
    pyeong = round(area / 3.3058, 1) if area > 0 else 0

    atcl_no = article.get("atclNo", "")
    link = f"https://fin.land.naver.com/articles/{atcl_no}" if atcl_no else ""

    return {
        "매물번호": atcl_no,
        "매물명": article.get("atclNm", "") or article.get("bildNm", ""),
        "유형": article.get("rletTpNm", ""),
        "매매가_만원": prc,
        "매매가_억": round(prc / 10000, 1) if prc >= 10000 else f"{prc}만",
        "면적_m2": round(area, 1),
        "면적_평": pyeong,
        "층": article.get("flrInfo", ""),
        "설명": article.get("atclFetrDesc", "") or "",
        "중개사": article.get("rltrNm", ""),
        "확인일": article.get("cfmYmd", ""),
        "링크": link,
    }


RECENT_DAYS = 1  # 기본: 어제 등록 매물만


def collect_buildings_for_gungu(gungu_info, log=print):
    """하나의 시군구에서 꼬마빌딩 매물 수집"""
    sido = gungu_info["sido"]
    gungu = gungu_info["gungu"]
    cortar_no = gungu_info["cortarNo"]
    lat = gungu_info["lat"]
    lon = gungu_info["lon"]

    buildings = []
    seen = set()

    for rlet_code in BUILDING_SEARCH_CODES:
        try:
            clusters = fetch_clusters(lat, lon, cortar_no, rlet_code)
        except Exception as e:
            log(f"  ⚠ {sido} {gungu} 클러스터 조회 실패 ({rlet_code}): {e}")
            continue
        wait()

        for cluster in clusters[:5]:  # 시군구당 최대 5개 클러스터만
            try:
                articles = fetch_articles_from_cluster(cluster, cortar_no, rlet_code)
            except Exception as e:
                continue
            wait()

            for a in articles:
                atcl_no = a.get("atclNo")
                if not atcl_no or atcl_no in seen:
                    continue
                seen.add(atcl_no)

                prc = int(a.get("prc", 0) or 0)
                if prc <= 0 or prc > MAX_PRICE:
                    continue

                if not is_building_type(a):
                    continue

                if not is_recent(a, days=RECENT_DAYS):
                    continue

                buildings.append(parse_article(a))

    return buildings


def collect_all(sido_filter=None, log=print):
    """서울/경기/인천 전체 수집"""
    if sido_filter:
        codes = {k: v for k, v in TARGET_SIDO.items() if k in sido_filter}
    else:
        codes = TARGET_SIDO

    all_gungu = get_all_gungu(codes, log)
    log(f"\n총 {len(all_gungu)}개 시군구 검색 시작\n")

    all_buildings = []
    for i, gungu in enumerate(all_gungu, 1):
        label = f"{gungu['sido']} {gungu['gungu']}"
        log(f"[{i}/{len(all_gungu)}] {label} 검색 중...")

        buildings = collect_buildings_for_gungu(gungu, log)
        if buildings:
            for b in buildings:
                b["지역_시도"] = gungu["sido"]
                b["지역_시군구"] = gungu["gungu"]
            all_buildings.extend(buildings)
            log(f"  → {len(buildings)}건 발견")
        else:
            log(f"  → 0건")

    log(f"\n총 {len(all_buildings)}건 수집 완료")
    return all_buildings


# ── 3. 신규 매물 판별 ────────────────────────────────

def load_previous_ids():
    """이전 수집 데이터에서 매물번호 목록 로드"""
    id_file = os.path.join(DATA_DIR, "collected_ids.json")
    if os.path.exists(id_file):
        with open(id_file, "r", encoding="utf-8") as f:
            return set(json.load(f))
    return set()


def save_current_ids(buildings):
    """현재 수집된 매물번호 저장"""
    os.makedirs(DATA_DIR, exist_ok=True)
    id_file = os.path.join(DATA_DIR, "collected_ids.json")

    # 기존 ID 로드 + 신규 추가
    existing = load_previous_ids()
    current = {b["매물번호"] for b in buildings}
    merged = existing | current

    with open(id_file, "w", encoding="utf-8") as f:
        json.dump(list(merged), f, ensure_ascii=False)

    return current - existing  # 신규 ID 반환


def filter_new_buildings(buildings):
    """이전에 없던 신규 매물만 필터링"""
    prev_ids = load_previous_ids()
    if not prev_ids:
        return buildings  # 첫 실행이면 전부 신규

    new_buildings = [b for b in buildings if b["매물번호"] not in prev_ids]
    return new_buildings


# ── 4. 엑셀 리포트 생성 ──────────────────────────────

def style_header(ws, headers, row=1):
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
    border = Border(*(Side(style="thin") for _ in range(4)))
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            if cell.value:
                length = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                max_len = max(max_len, length)
        ws.column_dimensions[letter].width = min(max_len + 4, 50)


def create_report(buildings, filename, title="전체"):
    """엑셀 리포트 생성"""
    wb = Workbook()
    link_font = Font(color="0563C1", underline="single")
    num_fmt = "#,##0"

    # -- 매물 목록 시트 --
    ws = wb.active
    ws.title = "매물목록"

    col_headers = [
        "지역", "매물명", "유형", "매매가(억)", "면적(m2)", "면적(평)",
        "층", "설명", "중개사", "네이버부동산", "링크(복사용)",
    ]
    style_header(ws, col_headers)

    # 가격순 정렬
    sorted_buildings = sorted(buildings, key=lambda x: x["매매가_만원"])

    for i, b in enumerate(sorted_buildings, 2):
        ws.cell(row=i, column=1, value=f"{b['지역_시도']} {b['지역_시군구']}")
        ws.cell(row=i, column=2, value=b["매물명"])
        ws.cell(row=i, column=3, value=b["유형"])
        ws.cell(row=i, column=4, value=b["매매가_억"])
        ws.cell(row=i, column=5, value=b["면적_m2"]).number_format = "#,##0.0"
        ws.cell(row=i, column=6, value=b["면적_평"]).number_format = "#,##0.0"
        ws.cell(row=i, column=7, value=b["층"])
        ws.cell(row=i, column=8, value=b["설명"])
        ws.cell(row=i, column=9, value=b["중개사"])

        url = b.get("링크", "")
        if url:
            cell = ws.cell(row=i, column=10, value="보기")
            cell.hyperlink = url
            cell.font = link_font
            ws.cell(row=i, column=11, value=url)

    auto_width(ws)

    # -- 요약 시트 --
    ws_sum = wb.create_sheet("요약")
    style_header(ws_sum, ["항목", "값"])

    summary = [
        ("조사 일시", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("검색 지역", title),
        ("총 매물 수", f"{len(buildings)}건"),
        ("가격 기준", "10억 이하"),
    ]

    # 지역별 통계
    region_counts = {}
    for b in buildings:
        key = f"{b['지역_시도']} {b['지역_시군구']}"
        region_counts[key] = region_counts.get(key, 0) + 1

    for region, cnt in sorted(region_counts.items(), key=lambda x: -x[1]):
        summary.append((region, f"{cnt}건"))

    # 가격 통계
    prices = [b["매매가_만원"] for b in buildings if b["매매가_만원"] > 0]
    if prices:
        avg = round(sum(prices) / len(prices))
        summary.append(("평균 매매가", f"{round(avg/10000, 1)}억원"))
        summary.append(("최저가", f"{round(min(prices)/10000, 1)}억원"))
        summary.append(("최고가", f"{round(max(prices)/10000, 1)}억원"))

    for i, (label, value) in enumerate(summary, 2):
        ws_sum.cell(row=i, column=1, value=label)
        ws_sum.cell(row=i, column=2, value=value)
    auto_width(ws_sum)

    wb.save(filename)
    return filename


# ── 5. 메인 실행 ─────────────────────────────────────

def main():
    global RECENT_DAYS

    # 인자 파싱
    region_filter = None
    if "--region" in sys.argv:
        idx = sys.argv.index("--region")
        if idx + 1 < len(sys.argv):
            region_filter = [sys.argv[idx + 1]]

    if "--days" in sys.argv:
        idx = sys.argv.index("--days")
        if idx + 1 < len(sys.argv):
            RECENT_DAYS = int(sys.argv[idx + 1])

    print("=" * 50)
    print("  소액 꼬마빌딩 급매물 수집기")
    print("  10억 이하 | 서울/경기/인천")
    print("=" * 50)
    print()

    # 수집
    buildings = collect_all(sido_filter=region_filter)

    if not buildings:
        print("\n매물이 없습니다.")
        if "--notify" in sys.argv:
            send_telegram("오늘 신규 꼬마빌딩 매물이 없습니다.")
        return

    # 신규 매물 판별
    new_buildings = filter_new_buildings(buildings)
    print(f"\n신규 매물: {len(new_buildings)}건 / 전체: {len(buildings)}건")

    # ID 저장
    save_current_ids(buildings)

    # 엑셀 생성 (전체)
    os.makedirs(DB_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    region_label = region_filter[0] if region_filter else "서울경기인천"

    filename = os.path.join(DB_DIR, f"급매물_{region_label}_{timestamp}.xlsx")
    create_report(buildings, filename, title=region_label)
    print(f"\n전체 리포트 저장: {filename}")

    # 신규 매물 엑셀 (있을 때만)
    if new_buildings and len(new_buildings) < len(buildings):
        new_filename = os.path.join(DB_DIR, f"신규매물_{region_label}_{timestamp}.xlsx")
        create_report(new_buildings, new_filename, title=f"{region_label} 신규")
        print(f"신규 리포트 저장: {new_filename}")

    # 텔레그램 알림
    if "--notify" in sys.argv:
        notify_buildings(buildings, excel_path=filename)
        print("텔레그램 알림 발송 완료")

    print(f"\n완료!")


if __name__ == "__main__":
    main()
