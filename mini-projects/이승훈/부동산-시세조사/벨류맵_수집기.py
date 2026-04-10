"""
벨류맵(valueupmap.com) 꼬마빌딩 급매물 수집기
- Playwright로 카카오 로그인 → 세션(쿠키) 저장
- 저장된 쿠키로 벨류맵 API 직접 호출 (빠르고 안정적)
- 서울 전 구 10억 이하 꼬마빌딩 매물 수집
- 시세대비 저평가 매물 자동 감지
- 텔레그램 알림 + 엑셀 리포트

사용법:
  python 벨류맵_수집기.py --login          # 첫 실행: 카카오 로그인 + 세션 저장
  python 벨류맵_수집기.py                   # 수집 실행
  python 벨류맵_수집기.py --notify          # 수집 + 텔레그램 알림
  python 벨류맵_수집기.py --debug           # 브라우저 보이면서 수집
"""

import os
import sys
import json
import time
import random
import requests as http_requests
from datetime import datetime
from urllib.parse import quote
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# .env 로드
load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))

# ── 설정 ──────────────────────────────────────────────

DELAY_MIN = 2.0
DELAY_MAX = 4.0

MAX_PRICE = 2000000000  # 원 단위 (20억)

# 저장 경로
DB_DIR = os.path.join(os.path.expanduser("~"), "OneDrive", "\ubc14\ud0d5 \ud654\uba74", "DB")
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(SCRIPT_DIR, "data")
SESSION_DIR = os.path.join(DATA_DIR, "valuemap_session")
SESSION_FILE = os.path.join(SESSION_DIR, "state.json")
IDS_FILE = os.path.join(DATA_DIR, "valuemap_ids.json")

# 카카오 로그인
KAKAO_ID = os.getenv("KAKAO_ID", "")
KAKAO_PW = os.getenv("KAKAO_PW", "")

# 텔레그램
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN", "")
TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID", "")

# 벨류맵 API
VALUEMAP_BASE = "https://www.valueupmap.com"
VALUEMAP_API = f"{VALUEMAP_BASE}/api"

# 지역 코드 (districtCode)
DISTRICTS = {
    # 서울
    "강남구": "1168000000", "강동구": "1174000000", "강북구": "1130500000",
    "강서구": "1150000000", "관악구": "1162000000", "광진구": "1121500000",
    "구로구": "1153000000", "금천구": "1154500000", "노원구": "1135000000",
    "도봉구": "1132000000", "동대문구": "1123000000", "동작구": "1159000000",
    "마포구": "1144000000", "서대문구": "1141000000", "서초구": "1165000000",
    "성동구": "1120000000", "성북구": "1129000000", "송파구": "1171000000",
    "양천구": "1147000000", "영등포구": "1156000000", "용산구": "1117000000",
    "은평구": "1138000000", "종로구": "1111000000", "중구(서울)": "1114000000",
    "중랑구": "1126000000",
    # 수원
    "장안구": "4111100000", "권선구": "4111300000",
    "팔달구": "4111500000", "영통구": "4111700000",
}


# ── 유틸 ──────────────────────────────────────────────

def wait():
    time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))


def log(msg):
    text = f"[{datetime.now().strftime('%H:%M:%S')}] {msg}"
    sys.stdout.buffer.write((text + "\n").encode("utf-8", errors="replace"))
    sys.stdout.flush()


# ── 1. 로그인 세션 관리 ──────────────────────────────

def login_and_save_session():
    """카카오 자동 로그인 -> 세션 저장"""
    from playwright.sync_api import sync_playwright

    if not KAKAO_ID or not KAKAO_PW:
        log("KAKAO_ID, KAKAO_PW가 .env에 설정되어 있지 않습니다.")
        sys.exit(1)

    os.makedirs(SESSION_DIR, exist_ok=True)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context(
            viewport={"width": 1280, "height": 900},
            locale="ko-KR",
        )

        popup_page = [None]
        context.on("page", lambda pg: popup_page.__setitem__(0, pg))

        page = context.new_page()

        login_url = "https://accounts.valueupmap.com/web/accounts/login?curl=https://www.valueupmap.com//me"
        log("벨류맵 로그인 페이지로 이동...")
        page.goto(login_url, timeout=60000, wait_until="domcontentloaded")
        page.wait_for_timeout(3000)

        log("카카오 로그인 버튼 클릭...")
        page.evaluate('''() => {
            const popup = document.querySelector('.popup--login');
            if (popup) popup.style.display = 'block';
        }''')
        page.wait_for_timeout(500)

        kakao_btn = page.query_selector('#kakaoLoginPopup')
        if not kakao_btn or not kakao_btn.is_visible():
            log("카카오 로그인 버튼을 찾지 못했습니다.")
            browser.close()
            return

        kakao_btn.click()
        page.wait_for_timeout(5000)

        kakao_page = popup_page[0]
        if not kakao_page:
            log("카카오 로그인 팝업이 열리지 않았습니다.")
            browser.close()
            return

        log("카카오 로그인 팝업 감지...")
        kakao_page.wait_for_load_state("domcontentloaded")
        page.wait_for_timeout(2000)

        log("카카오 계정 입력 중...")
        try:
            id_input = kakao_page.wait_for_selector(
                'input[name="loginId"], input#loginId', timeout=15000,
            )
            id_input.fill(KAKAO_ID)
            page.wait_for_timeout(1000)

            pw_input = kakao_page.query_selector(
                'input[name="password"], input#password, input[type="password"]'
            )
            if pw_input:
                pw_input.fill(KAKAO_PW)
            page.wait_for_timeout(1000)

            submit_btn = kakao_page.query_selector(
                'button[type="submit"], button.submit, button:has-text("로그인")'
            )
            if submit_btn:
                submit_btn.click()
            else:
                kakao_page.keyboard.press("Enter")

            log("로그인 시도 중...")
            page.wait_for_timeout(10000)

        except Exception as e:
            log(f"카카오 로그인 실패: {e}")
            browser.close()
            return

        log(f"메인 페이지 URL: {page.url}")

        page.goto(VALUEMAP_BASE, timeout=60000, wait_until="domcontentloaded")
        page.wait_for_timeout(5000)

        context.storage_state(path=SESSION_FILE)
        log(f"세션 저장 완료")

        browser.close()

    log("다음부터는 자동 로그인으로 실행됩니다.")


def load_api_session():
    """저장된 세션 쿠키를 requests.Session에 로드"""
    if not os.path.exists(SESSION_FILE):
        log("저장된 세션이 없습니다. --login으로 먼저 로그인하세요.")
        sys.exit(1)

    with open(SESSION_FILE, "r", encoding="utf-8") as f:
        state = json.load(f)

    session = http_requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Referer": "https://www.valueupmap.com/map",
        "Accept": "application/json",
    })

    for cookie in state.get("cookies", []):
        session.cookies.set(cookie["name"], cookie["value"], domain=cookie.get("domain", ""))

    return session


def check_session(session):
    """API 세션이 유효한지 확인"""
    resp = session.get(
        f"{VALUEMAP_API}/accounts/notifications/me/unread-count",
        timeout=10,
    )
    return resp.status_code == 200


# ── 2. 매물 수집 (API) ──────────────────────────────

def fetch_items_for_district(session, district_name, district_code, log_fn=log):
    """한 구의 매물을 API로 수집 (페이지네이션)"""
    items = []
    page_num = 0
    page_size = 50

    while True:
        resp = session.get(
            f"{VALUEMAP_API}/plus/item-groups",
            params={
                "areaUnitType": "PYEONG",
                "dealTypes": "SALES",
                "districtCode": district_code,
                "isSearched": "false",
                "isSelected": "true",
                "page": str(page_num),
                "size": str(page_size),
                "sort": "DEFAULT",
            },
            timeout=15,
        )

        if resp.status_code == 401:
            log_fn("세션 만료! --login으로 다시 로그인하세요.")
            return None

        if resp.status_code != 200:
            log_fn(f"  {district_name}: API 오류 ({resp.status_code})")
            break

        data = resp.json()
        contents = data.get("contents", [])
        total = data.get("count", 0)

        for item in contents:
            price = item.get("price", 0) or 0
            if price > MAX_PRICE or price <= 0:
                continue

            prop_type = item.get("propertyType1", {}).get("code", "")
            if prop_type != "BUILDING":
                continue

            parsed = parse_item(item, district_name)
            items.append(parsed)

        if not contents or (page_num + 1) * page_size >= total:
            break

        page_num += 1
        wait()

    return items


def parse_item(item, district_name):
    """API 응답 아이템을 정리"""
    price = item.get("price", 0) or 0
    item_id = str(item.get("id", ""))

    attrs = {}
    for a in item.get("summaryAttributes", []):
        attrs[a["title"]] = a["content"]

    land_area = float(attrs.get("대지면적", "0") or "0")
    building_area = float(attrs.get("연면적", "0") or "0")
    land_pyeong = round(land_area / 3.3058, 1) if land_area > 0 else 0
    building_pyeong = round(building_area / 3.3058, 1) if building_area > 0 else 0

    land_price_per_pyeong = item.get("landPricePerPyeong", 0) or 0
    building_price_per_pyeong = item.get("buildingPricePerPyeong", 0) or 0

    mps = item.get("marketPriceStatus")
    market_status = mps.get("label", "") if mps else ""

    return {
        "매물번호": item_id,
        "지역": district_name,
        "주소": item.get("shortAddress", ""),
        "유형": item.get("propertyType2", {}).get("label", ""),
        "거래방식": item.get("dealMethodType", {}).get("label", "") if item.get("dealMethodType") else "",
        "매매가_원": price,
        "매매가_억": round(price / 100000000, 1),
        "대지면적_m2": land_area,
        "대지면적_평": land_pyeong,
        "연면적_m2": building_area,
        "연면적_평": building_pyeong,
        "토지_평단가": round(land_price_per_pyeong / 10000) if land_price_per_pyeong else 0,
        "건물_평단가": round(building_price_per_pyeong / 10000) if building_price_per_pyeong else 0,
        "규모": attrs.get("규모", ""),
        "노후도": attrs.get("노후도", ""),
        "시세판단": market_status,
        "조회수": item.get("hitCount", 0),
        "링크": f"{VALUEMAP_BASE}/properties/rental-items/{item.get('pnu', '')}?address={quote(item.get('shortAddress', ''))}&dealType=SALES",
    }


def collect_all(session, log_fn=log):
    """서울+수원 전 구 수집"""
    all_buildings = []

    for district_name, district_code in DISTRICTS.items():
        log_fn(f"  {district_name} 수집 중...")
        items = fetch_items_for_district(session, district_name, district_code, log_fn)

        if items is None:
            return None  # 세션 만료

        if items:
            log_fn(f"    -> {len(items)}건")
        all_buildings.extend(items)
        wait()

    log_fn(f"\n총 {len(all_buildings)}건 수집 완료")
    return all_buildings


# ── 3. 신규 매물 판별 ───────────────────────────────

def load_previous_ids():
    if os.path.exists(IDS_FILE):
        with open(IDS_FILE, "r", encoding="utf-8") as f:
            return set(json.load(f))
    return set()


def save_current_ids(buildings):
    os.makedirs(DATA_DIR, exist_ok=True)
    existing = load_previous_ids()
    current = {b["매물번호"] for b in buildings}
    merged = existing | current
    with open(IDS_FILE, "w", encoding="utf-8") as f:
        json.dump(list(merged), f, ensure_ascii=False)
    return current - existing


def is_first_run():
    if not os.path.exists(IDS_FILE):
        return True
    with open(IDS_FILE, "r", encoding="utf-8") as f:
        ids = json.load(f)
    return len(ids) == 0


def filter_new_buildings(buildings):
    prev_ids = load_previous_ids()
    if not prev_ids:
        return []
    return [b for b in buildings if b["매물번호"] not in prev_ids]


# ── 4. 엑셀 리포트 ──────────────────────────────────

def style_header(ws, headers, row=1):
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
    border = Border(*(Side(style="thin") for _ in range(4)))
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            if cell.value:
                length = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                max_len = max(max_len, length)
        ws.column_dimensions[letter].width = min(max_len + 4, 50)


def create_report(buildings, filename, title="서울+수원"):
    wb = Workbook()
    link_font = Font(color="0563C1", underline="single")
    num_fmt = "#,##0"

    ws = wb.active
    ws.title = "매물목록"

    col_headers = [
        "지역", "주소", "유형", "거래방식", "매매가(억)",
        "대지면적(m2)", "대지(평)", "연면적(m2)", "연면적(평)",
        "토지평단가(만)", "건물평단가(만)", "규모", "노후도",
        "시세판단", "조회수", "벨류맵", "링크(복사용)",
    ]
    style_header(ws, col_headers)

    # 가격 오름차순 정렬
    sorted_buildings = sorted(buildings, key=lambda x: x["매매가_원"])

    # 시세대비 저평가 강조
    below_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")

    for i, b in enumerate(sorted_buildings, 2):
        ws.cell(row=i, column=1, value=b["지역"])
        ws.cell(row=i, column=2, value=b["주소"])
        ws.cell(row=i, column=3, value=b["유형"])
        ws.cell(row=i, column=4, value=b["거래방식"])
        ws.cell(row=i, column=5, value=b["매매가_억"])
        ws.cell(row=i, column=6, value=b["대지면적_m2"]).number_format = "#,##0.0"
        ws.cell(row=i, column=7, value=b["대지면적_평"]).number_format = "#,##0.0"
        ws.cell(row=i, column=8, value=b["연면적_m2"]).number_format = "#,##0.0"
        ws.cell(row=i, column=9, value=b["연면적_평"]).number_format = "#,##0.0"
        ws.cell(row=i, column=10, value=b["토지_평단가"]).number_format = num_fmt
        ws.cell(row=i, column=11, value=b["건물_평단가"]).number_format = num_fmt
        ws.cell(row=i, column=12, value=b["규모"])
        ws.cell(row=i, column=13, value=b["노후도"])
        ws.cell(row=i, column=14, value=b["시세판단"])
        ws.cell(row=i, column=15, value=b["조회수"]).number_format = num_fmt

        url = b.get("링크", "")
        if url:
            cell = ws.cell(row=i, column=16, value="보기")
            cell.hyperlink = url
            cell.font = link_font
            ws.cell(row=i, column=17, value=url)

        # 저평가 매물 하이라이트
        if "저" in b.get("시세판단", "") or "하" in b.get("시세판단", ""):
            for col in range(1, 18):
                ws.cell(row=i, column=col).fill = below_fill

    auto_width(ws)

    # 요약 시트
    ws_sum = wb.create_sheet("요약")
    style_header(ws_sum, ["항목", "값"])
    summary = [
        ("조사 일시", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("데이터 소스", "벨류맵 (valueupmap.com)"),
        ("검색 지역", title),
        ("총 매물 수", f"{len(buildings)}건"),
        ("가격 기준", "20억 이하"),
    ]

    # 구별 통계
    region_counts = {}
    for b in buildings:
        region_counts[b["지역"]] = region_counts.get(b["지역"], 0) + 1
    for region, cnt in sorted(region_counts.items(), key=lambda x: -x[1]):
        summary.append((region, f"{cnt}건"))

    # 가격 통계
    prices = [b["매매가_원"] for b in buildings if b["매매가_원"] > 0]
    if prices:
        summary.append(("평균 매매가", f"{round(sum(prices)/len(prices)/100000000, 1)}억원"))
        summary.append(("최저가", f"{round(min(prices)/100000000, 1)}억원"))
        summary.append(("최고가", f"{round(max(prices)/100000000, 1)}억원"))

    # 시세판단 분포
    status_counts = {}
    for b in buildings:
        s = b.get("시세판단", "미분류")
        status_counts[s] = status_counts.get(s, 0) + 1
    for s, cnt in sorted(status_counts.items(), key=lambda x: -x[1]):
        summary.append((f"시세판단: {s}", f"{cnt}건"))

    for i, (label, value) in enumerate(summary, 2):
        ws_sum.cell(row=i, column=1, value=label)
        ws_sum.cell(row=i, column=2, value=value)
    auto_width(ws_sum)

    wb.save(filename)
    return filename


# ── 5. 텔레그램 알림 ────────────────────────────────

def send_telegram(text):
    if not TELEGRAM_TOKEN or not TELEGRAM_CHAT_ID:
        return
    try:
        http_requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
            json={
                "chat_id": TELEGRAM_CHAT_ID,
                "text": text,
                "parse_mode": "HTML",
                "disable_web_page_preview": True,
            },
            timeout=10,
        )
    except Exception:
        pass


def send_telegram_file(filepath, caption=""):
    if not TELEGRAM_TOKEN or not TELEGRAM_CHAT_ID:
        return
    try:
        with open(filepath, "rb") as f:
            http_requests.post(
                f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendDocument",
                data={"chat_id": TELEGRAM_CHAT_ID, "caption": caption},
                files={"document": f},
                timeout=30,
            )
    except Exception:
        pass


# ── 메인 ──────────────────────────────────────────────

def main():
    do_login = "--login" in sys.argv
    notify = "--notify" in sys.argv

    log("=" * 50)
    log("  벨류맵 꼬마빌딩 급매물 수집기")
    log("  서울+수원 | 20억 이하 | 건물(빌딩)")
    log("=" * 50)

    if do_login:
        login_and_save_session()
        return

    # API 세션 로드
    session = load_api_session()

    log("세션 확인 중...")
    if not check_session(session):
        log("세션이 만료되었습니다. --login으로 다시 로그인해주세요.")
        if notify:
            send_telegram("[벨류맵] 세션 만료 - 재로그인 필요")
        sys.exit(1)
    log("세션 유효")

    # 서울 전 구 매물 수집
    log("\n서울+수원 매물 수집 시작...")
    buildings = collect_all(session, log_fn=log)

    if buildings is None:
        log("세션 만료로 수집 중단. --login으로 다시 로그인하세요.")
        if notify:
            send_telegram("[벨류맵] 세션 만료 - 재로그인 필요")
        sys.exit(1)

    if not buildings:
        log("수집된 매물이 없습니다.")
        if notify:
            send_telegram("[벨류맵] 수집된 매물이 없습니다.")
        return

    # 신규 매물 판별
    first_run = is_first_run()
    new_buildings = filter_new_buildings(buildings)
    save_current_ids(buildings)

    if first_run:
        log(f"\n[기준선 수집] 전체 {len(buildings)}건 등록 완료")
        log("  다음 실행부터 신규 매물만 따로 분류됩니다.")
    else:
        log(f"\n신규 매물: {len(new_buildings)}건 / 전체: {len(buildings)}건")

    # 엑셀 리포트
    os.makedirs(DB_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")

    filename = os.path.join(DB_DIR, f"VM_급매물_{timestamp}.xlsx")
    create_report(buildings, filename)
    log(f"전체 리포트 저장: {filename}")

    if not first_run and new_buildings:
        new_filename = os.path.join(DB_DIR, f"VM_신규매물_{timestamp}.xlsx")
        create_report(new_buildings, new_filename, title="서울 신규")
        log(f"신규 리포트 저장: {new_filename}")

    # 텔레그램 알림
    if notify:
        if first_run:
            send_telegram(f"[벨류맵] 기준선 수집 완료: {len(buildings)}건")
        else:
            if new_buildings:
                send_telegram_file(filename, caption=f"[벨류맵] 신규 매물 {len(new_buildings)}건")
            else:
                send_telegram("[벨류맵] 오늘 신규 매물 없음")
        log("텔레그램 알림 발송 완료")

    log("\n완료!")


if __name__ == "__main__":
    main()
