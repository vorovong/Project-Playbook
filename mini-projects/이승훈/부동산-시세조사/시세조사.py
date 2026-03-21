"""
네이버 부동산 상가 시세조사 자동화
- GUI: python 시세조사.py
- CLI: python 시세조사.py 팔달구 장안동
"""

import os
import sys
import time
import threading
import tkinter as tk
from tkinter import ttk, messagebox
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ── 설정 ──────────────────────────────────────────────

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Referer": "https://m.land.naver.com/",
}

DELAY = 1.0

TRADE_TYPES = {
    "매매": "A1",
    "전세": "B1",
    "월세": "B2",
}

DESKTOP = os.path.join(os.path.expanduser("~"), "Desktop")


# ── 1. 동 이름 → cortarNo 변환 ───────────────────────

def get_regions(cortar_no):
    resp = requests.get(
        "https://m.land.naver.com/map/getRegionList",
        params={"cortarNo": cortar_no},
        headers=HEADERS,
        timeout=10,
    )
    resp.raise_for_status()
    data = resp.json()
    result = data.get("result", {})
    return result.get("list", []), result.get("dvsnInfo", {}), result.get("cityInfo", {})


def find_dong(query_parts, log=print):
    dong_name = query_parts[-1]
    filters = query_parts[:-1]
    matches = []

    sido_list, _, _ = get_regions("0000000000")
    time.sleep(DELAY)

    for sido in sido_list:
        sido_nm = sido["CortarNm"]
        log(f"  검색 중: {sido_nm}...")

        gungu_list, _, _ = get_regions(sido["CortarNo"])
        time.sleep(DELAY)

        for gungu in gungu_list:
            gungu_nm = gungu["CortarNm"]
            full_path = f"{sido_nm} {gungu_nm}"

            if filters and not all(f in full_path for f in filters):
                continue

            dong_list, _, _ = get_regions(gungu["CortarNo"])
            time.sleep(DELAY)

            for dong in dong_list:
                if dong_name in dong["CortarNm"]:
                    full_name = f"{sido_nm} {gungu_nm} {dong['CortarNm']}"
                    matches.append({
                        "name": full_name,
                        "cortarNo": dong["CortarNo"],
                        "lat": float(dong["MapYCrdn"]),
                        "lon": float(dong["MapXCrdn"]),
                    })

        if not filters and matches:
            break

    return matches


# ── 2. 매물 조회 ─────────────────────────────────────

def fetch_clusters(lat, lon, cortar_no, trade_type, rlet_type="D02"):
    offset = 0.015
    resp = requests.get(
        "https://m.land.naver.com/cluster/clusterList",
        params={
            "view": "atcl", "rletTpCd": rlet_type, "tradTpCd": trade_type,
            "z": "16", "lat": str(lat), "lon": str(lon),
            "btm": str(lat - offset), "lft": str(lon - offset),
            "top": str(lat + offset), "rgt": str(lon + offset),
            "cortarNo": cortar_no,
        },
        headers=HEADERS, timeout=10,
    )
    resp.raise_for_status()
    return resp.json().get("data", {}).get("ARTICLE", [])


def fetch_articles_from_cluster(cluster, cortar_no, trade_type, rlet_type="D02"):
    all_articles = []
    page = 1
    offset = 0.015
    lat, lon = cluster["lat"], cluster["lon"]

    while True:
        resp = requests.get(
            "https://m.land.naver.com/cluster/ajax/articleList",
            params={
                "rletTpCd": rlet_type, "tradTpCd": trade_type, "z": "16",
                "lat": str(lat), "lon": str(lon),
                "btm": str(lat - offset), "lft": str(lon - offset),
                "top": str(lat + offset), "rgt": str(lon + offset),
                "lgeo": cluster["lgeo"], "showR0": "N",
                "page": str(page), "cortarNo": cortar_no,
            },
            headers=HEADERS, timeout=10,
        )
        resp.raise_for_status()
        data = resp.json()
        body = data.get("body", [])
        if not body:
            break
        all_articles.extend(body)
        if not data.get("more", False):
            break
        page += 1
        time.sleep(DELAY)

    return all_articles


def fetch_all_articles(lat, lon, cortar_no, trade_type, rlet_type="D02"):
    clusters = fetch_clusters(lat, lon, cortar_no, trade_type, rlet_type)
    time.sleep(DELAY)
    all_articles = []
    seen = set()
    for cluster in clusters:
        for a in fetch_articles_from_cluster(cluster, cortar_no, trade_type, rlet_type):
            atcl_no = a.get("atclNo")
            if atcl_no and atcl_no not in seen:
                seen.add(atcl_no)
                all_articles.append(a)
        time.sleep(DELAY)
    return all_articles


# ── 3. 주소 변환 / 파싱 ──────────────────────────────

COMMERCIAL_TYPES = {"상가", "사무실", "상가주택", "건물", "공장/창고", "토지"}


def reverse_geocode_batch(articles, log=print):
    addr_cache = {}
    coords_needed = set()
    for a in articles:
        lat = round(float(a.get("lat", 0) or 0), 4)
        lng = round(float(a.get("lng", 0) or 0), 4)
        if lat and lng:
            coords_needed.add((lat, lng))

    log(f"  주소 변환 중... ({len(coords_needed)}곳)")
    for lat, lng in coords_needed:
        try:
            resp = requests.get(
                "https://nominatim.openstreetmap.org/reverse",
                params={
                    "lat": str(lat), "lon": str(lng), "format": "json",
                    "addressdetails": "1", "accept-language": "ko", "zoom": "18",
                },
                headers={"User-Agent": "real-estate-survey/1.0"}, timeout=10,
            )
            if resp.status_code == 200:
                addr = resp.json().get("address", {})
                road = addr.get("road", "")
                house = addr.get("house_number", "")
                borough = addr.get("borough", addr.get("suburb", ""))
                city = addr.get("city", addr.get("town", ""))
                parts = [p for p in [city, borough, road, house] if p]
                addr_cache[(lat, lng)] = " ".join(parts)
        except Exception:
            pass
        time.sleep(1.1)
    return addr_cache


def parse_articles(articles, trade_type, log=print):
    commercial = [a for a in articles if a.get("rletTpNm", "") in COMMERCIAL_TYPES]
    addr_cache = reverse_geocode_batch(commercial, log)

    result = []
    for a in commercial:
        try:
            spc1 = float(a.get("spc1", 0) or 0)
        except (ValueError, TypeError):
            spc1 = 0
        try:
            spc2 = float(a.get("spc2", 0) or 0)
        except (ValueError, TypeError):
            spc2 = 0
        area = spc2 if spc2 > 0 else spc1
        pyeong = round(area / 3.3058, 1) if area > 0 else 0

        lat = round(float(a.get("lat", 0) or 0), 4)
        lng = round(float(a.get("lng", 0) or 0), 4)
        address = addr_cache.get((lat, lng), "")

        atcl_no = a.get("atclNo", "")
        link = f"https://fin.land.naver.com/articles/{atcl_no}" if atcl_no else ""

        item = {
            "매물명": a.get("atclNm", "") or a.get("bildNm", ""),
            "유형": a.get("rletTpNm", ""),
            "주소": address,
            "면적_m2": round(area, 1),
            "면적_평": pyeong,
            "층": a.get("flrInfo", ""),
            "설명": a.get("atclFetrDesc", ""),
            "중개사": a.get("rltrNm", ""),
            "링크": link,
        }

        if trade_type == "A1":
            prc = int(a.get("prc", 0) or 0)
            item["매매가_만원"] = prc
            item["평단가_만원"] = round(prc / pyeong) if pyeong > 0 else 0
        elif trade_type == "B1":
            item["전세가_만원"] = int(a.get("prc", 0) or 0)
        else:
            item["보증금_만원"] = int(a.get("prc", 0) or 0)
            item["월세_만원"] = int(a.get("rentPrc", 0) or 0)

        result.append(item)
    return result


# ── 4. 엑셀 생성 ─────────────────────────────────────

def style_header(ws, col_headers, row=1):
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    border = Border(*(Side(style="thin") for _ in range(4)))
    for col, h in enumerate(col_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            if cell.value:
                length = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                max_len = max(max_len, length)
        ws.column_dimensions[letter].width = min(max_len + 4, 50)


def create_excel(dong_info, datasets, filename):
    """
    datasets: dict of {trade_label: (trade_code, articles_list)}
    """
    wb = Workbook()
    num_fmt = "#,##0"
    first_sheet = True

    summary_rows = [
        ("조사 지역", dong_info["name"]),
        ("조사 일시", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]

    for label, (code, data) in datasets.items():
        if first_sheet:
            ws = wb.active
            ws.title = label
            first_sheet = False
        else:
            ws = wb.create_sheet(label)

        link_font = Font(color="0563C1", underline="single")

        if code == "A1":
            headers = ["매물명", "유형", "주소", "면적(m2)", "면적(평)", "층", "매매가(만원)", "평단가(만원/평)", "설명", "네이버부동산", "링크(복사용)"]
        elif code == "B1":
            headers = ["매물명", "유형", "주소", "면적(m2)", "면적(평)", "층", "전세가(만원)", "설명", "네이버부동산", "링크(복사용)"]
        else:
            headers = ["매물명", "유형", "주소", "면적(m2)", "면적(평)", "층", "보증금(만원)", "월세(만원)", "설명", "네이버부동산", "링크(복사용)"]

        style_header(ws, headers)
        link_col = len(headers) - 1  # 하이퍼링크 컬럼
        url_col = len(headers)       # URL 텍스트 컬럼

        for i, item in enumerate(data, 2):
            ws.cell(row=i, column=1, value=item["매물명"])
            ws.cell(row=i, column=2, value=item["유형"])
            ws.cell(row=i, column=3, value=item["주소"])
            ws.cell(row=i, column=4, value=item["면적_m2"]).number_format = "#,##0.0"
            ws.cell(row=i, column=5, value=item["면적_평"]).number_format = "#,##0.0"
            ws.cell(row=i, column=6, value=item["층"])

            if code == "A1":
                ws.cell(row=i, column=7, value=item["매매가_만원"]).number_format = num_fmt
                ws.cell(row=i, column=8, value=item["평단가_만원"]).number_format = num_fmt
                ws.cell(row=i, column=9, value=item["설명"])
            elif code == "B1":
                ws.cell(row=i, column=7, value=item["전세가_만원"]).number_format = num_fmt
                ws.cell(row=i, column=8, value=item["설명"])
            else:
                ws.cell(row=i, column=7, value=item["보증금_만원"]).number_format = num_fmt
                ws.cell(row=i, column=8, value=item["월세_만원"]).number_format = num_fmt
                ws.cell(row=i, column=9, value=item["설명"])

            # 하이퍼링크 + URL 텍스트
            url = item.get("링크", "")
            if url:
                cell = ws.cell(row=i, column=link_col, value="보기")
                cell.hyperlink = url
                cell.font = link_font
                ws.cell(row=i, column=url_col, value=url)

        auto_width(ws)
        summary_rows.append((f"{label} 매물 수", f"{len(data)}건"))

        # 통계
        if code == "A1" and data:
            prices = [d["평단가_만원"] for d in data if d["평단가_만원"] > 0]
            if prices:
                summary_rows.append(("매매 평단가 평균", f"{round(sum(prices)/len(prices)):,}만원/평"))
        elif code == "B1" and data:
            prices = [d["전세가_만원"] for d in data if d["전세가_만원"] > 0]
            if prices:
                summary_rows.append(("전세가 평균", f"{round(sum(prices)/len(prices)):,}만원"))
        elif code == "B2" and data:
            deps = [d["보증금_만원"] for d in data if d["보증금_만원"] > 0]
            rents = [d["월세_만원"] for d in data if d["월세_만원"] > 0]
            if deps and rents:
                summary_rows.append(("월세 평균", f"보증금 {round(sum(deps)/len(deps)):,} / 월 {round(sum(rents)/len(rents)):,}"))

    # 요약 시트
    ws_sum = wb.create_sheet("요약")
    style_header(ws_sum, ["항목", "값"])
    for i, (label, value) in enumerate(summary_rows, 2):
        ws_sum.cell(row=i, column=1, value=label)
        ws_sum.cell(row=i, column=2, value=value)
    auto_width(ws_sum)

    wb.save(filename)
    return filename


# ── 5. GUI ────────────────────────────────────────────

class App:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("부동산 시세조사")
        self.root.resizable(False, False)

        # 프레임
        frm = ttk.Frame(self.root, padding=20)
        frm.grid()

        # 지역 입력
        ttk.Label(frm, text="지역명", font=("맑은 고딕", 11)).grid(row=0, column=0, sticky="w")
        self.entry = ttk.Entry(frm, width=30, font=("맑은 고딕", 11))
        self.entry.grid(row=0, column=1, columnspan=3, padx=(10, 0), pady=5)
        self.entry.insert(0, "팔달구 장안동")

        # 거래 유형 체크박스
        ttk.Label(frm, text="거래유형", font=("맑은 고딕", 11)).grid(row=1, column=0, sticky="w", pady=(10, 0))
        self.chk_vars = {}
        for i, label in enumerate(TRADE_TYPES.keys()):
            var = tk.BooleanVar(value=(label == "월세"))
            chk = ttk.Checkbutton(frm, text=label, variable=var)
            chk.grid(row=1, column=i + 1, sticky="w", pady=(10, 0))
            self.chk_vars[label] = var

        # 조회 버튼
        self.btn = ttk.Button(frm, text="조회", command=self.on_search)
        self.btn.grid(row=2, column=0, columnspan=4, pady=(15, 5), sticky="ew")

        # 로그
        self.log_text = tk.Text(frm, width=55, height=15, font=("Consolas", 10), state="disabled")
        self.log_text.grid(row=3, column=0, columnspan=4, pady=(5, 0))

        self.root.mainloop()

    def log(self, msg):
        self.log_text.configure(state="normal")
        self.log_text.insert("end", msg + "\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")
        self.root.update_idletasks()

    def on_search(self):
        query = self.entry.get().strip()
        if not query:
            messagebox.showwarning("입력 필요", "지역명을 입력하세요.")
            return

        selected = [label for label, var in self.chk_vars.items() if var.get()]
        if not selected:
            messagebox.showwarning("선택 필요", "거래유형을 하나 이상 선택하세요.")
            return

        self.btn.configure(state="disabled")
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.configure(state="disabled")

        threading.Thread(target=self.run_search, args=(query, selected), daemon=True).start()

    def run_search(self, query, selected_types):
        try:
            query_parts = query.split()
            self.log(f"'{query}' 시세조사 시작\n")

            # 1. 동 찾기
            self.log("[1] 지역 검색 중...")
            matches = find_dong(query_parts, log=self.log)

            if not matches:
                self.log(f"\n'{query}' 검색 결과가 없습니다.")
                return

            if len(matches) > 1:
                # GUI에서 선택
                self.log(f"\n'{query}' 검색 결과:")
                for i, m in enumerate(matches, 1):
                    self.log(f"  {i}. {m['name']}")
                self.log("\n첫 번째 결과로 진행합니다.")

            dong_info = matches[0]
            self.log(f"\n-> {dong_info['name']}\n")

            lat = dong_info["lat"]
            lon = dong_info["lon"]
            cortar_no = dong_info["cortarNo"]

            # 2. 선택된 유형별 매물 조회
            datasets = {}
            step = 2
            for label in selected_types:
                code = TRADE_TYPES[label]
                self.log(f"[{step}] {label} 매물 조회 중...")
                raw = fetch_all_articles(lat, lon, cortar_no, code)
                data = parse_articles(raw, code, log=self.log)
                self.log(f"  -> {label} {len(data)}건\n")
                datasets[label] = (code, data)
                step += 1
                time.sleep(DELAY)

            # 3. 엑셀 저장
            self.log(f"[{step}] 엑셀 생성 중...")
            safe_name = dong_info["name"].replace(" ", "_")
            timestamp = datetime.now().strftime("%Y%m%d_%H%M")
            filename = os.path.join(DESKTOP, f"시세조사_{safe_name}_{timestamp}.xlsx")
            create_excel(dong_info, datasets, filename)

            self.log(f"\n완료! 바탕화면에 저장됨:")
            self.log(f"  {os.path.basename(filename)}")
            total = sum(len(d) for _, d in datasets.values())
            self.log(f"  총 {total}건")

            # 파일 열기
            os.startfile(filename)

        except Exception as e:
            self.log(f"\n오류 발생: {e}")
        finally:
            self.root.after(0, lambda: self.btn.configure(state="normal"))


# ── 메인 ──────────────────────────────────────────────

def main_cli():
    query_parts = sys.argv[1:]
    query_str = " ".join(query_parts)
    print(f"'{query_str}' 시세조사를 시작합니다...\n")

    print("[1] 지역 검색 중...")
    matches = find_dong(query_parts)
    if not matches:
        print(f"'{query_str}' 검색 결과가 없습니다.")
        sys.exit(1)
    if len(matches) == 1:
        dong_info = matches[0]
        print(f"  -> {dong_info['name']}")
    else:
        print(f"\n'{query_str}' 검색 결과:")
        for i, m in enumerate(matches, 1):
            print(f"  {i}. {m['name']}")
        while True:
            try:
                choice = int(input("번호 선택: "))
                if 1 <= choice <= len(matches):
                    dong_info = matches[choice - 1]
                    break
            except (ValueError, EOFError):
                pass

    lat, lon, cortar_no = dong_info["lat"], dong_info["lon"], dong_info["cortarNo"]

    datasets = {}
    for label, code in TRADE_TYPES.items():
        print(f"\n{label} 매물 조회 중...")
        raw = fetch_all_articles(lat, lon, cortar_no, code)
        data = parse_articles(raw, code)
        print(f"  -> {label} {len(data)}건")
        datasets[label] = (code, data)
        time.sleep(DELAY)

    safe_name = dong_info["name"].replace(" ", "_")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = os.path.join(DESKTOP, f"시세조사_{safe_name}_{timestamp}.xlsx")
    create_excel(dong_info, datasets, filename)
    print(f"\n완료! 파일: {filename}")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        main_cli()
    else:
        App()
