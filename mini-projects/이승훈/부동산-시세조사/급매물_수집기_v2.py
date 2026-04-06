"""
소액 꼬마빌딩 급매물 수집기 v2
- new.land.naver.com PC API 사용 (빠르고 안정적)
- 서울/경기/인천 10억 이하 상가건물 신규 매물 자동 수집
- 텔레그램 알림 + 엑셀 리포트

사용법:
  python 급매물_수집기_v2.py                    # 어제 신규, 서울/경기/인천
  python 급매물_수집기_v2.py --region 서울       # 서울만
  python 급매물_수집기_v2.py --days 3            # 최근 3일치
  python 급매물_수집기_v2.py --notify            # 텔레그램 알림 포함
"""

import os
import sys
import json
import time
import random
import requests
from datetime import datetime, timedelta
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# .env 로드 (스크립트 위치 기준)
load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))

# ── 설정 ──────────────────────────────────────────────

DELAY = 15.0  # API 호출 간 딜레이 (초)

# 조회 대상 시도 코드 (서울만)
TARGET_SIDO = {
    "서울": "1100000000",
}

# 꼬마빌딩 유형 — 모바일 API의 rletTpCd
# D02: 상가, D03: 사무실, E04: 단독/다가구
BUILDING_TYPES = "D02:D03:E04"

# 매물 상세에서 필터링할 유형명 — 꼬마빌딩(건물 통째)만
BUILDING_TYPE_NAMES = {"건물", "상가건물", "상가주택"}

# 가격 상한 (만원 단위, 10억 = 100000만원)
MAX_PRICE = 100000

# 날짜 필터 (기본: 어제 등록분만)
RECENT_DAYS = 1

# 저장 경로
DB_DIR = os.path.join(os.path.expanduser("~"), "OneDrive", "바탕 화면", "DB")
DATA_DIR = os.path.join(
    os.path.expanduser("~"),
    "projects", "Project_Playbook",
    "mini-projects", "이승훈", "부동산-시세조사", "data",
)

# 텔레그램 알림 설정 (.env에서 로드)
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN", "")
TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID", "")


# ── API 호출 ─────────────────────────────────────────

SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Referer": "https://m.land.naver.com/",
})


def api_get(url, params, retry=0):
    """API 호출 (재시도 포함). 429 시 None 반환으로 호출자에게 알림."""
    try:
        resp = SESSION.get(url, params=params, timeout=15)
        if resp.status_code == 429:
            if retry < 5:
                wait_time = (retry + 1) * 120
                print(f"  (429 - {wait_time}초 대기 후 재시도...)")
                time.sleep(wait_time)
                return api_get(url, params, retry + 1)
            print("  (429 - 재시도 초과, 스킵)")
            return None  # None = 429로 스킵됨
        resp.raise_for_status()
        return resp.json()
    except requests.exceptions.JSONDecodeError:
        if retry < 2:
            time.sleep(5)
            return api_get(url, params, retry + 1)
        return {}
    except Exception as e:
        print(f"  API 오류: {e}")
        return {}


def wait():
    """랜덤 딜레이"""
    time.sleep(DELAY + random.uniform(2, 5))


# ── 매물 수집 (모바일 API — 클러스터 방식) ──────────

def fetch_clusters(lat, lon, cortar_no):
    """m.land.naver.com 클러스터 목록 조회"""
    offset = 0.06
    data = api_get(
        "https://m.land.naver.com/cluster/clusterList",
        params={
            "view": "atcl", "rletTpCd": BUILDING_TYPES, "tradTpCd": "A1",
            "z": "13", "lat": str(lat), "lon": str(lon),
            "btm": str(lat - offset), "lft": str(lon - offset),
            "top": str(lat + offset), "rgt": str(lon + offset),
            "cortarNo": cortar_no,
        },
    )
    if data is None:
        return None
    return data.get("data", {}).get("ARTICLE", [])


def fetch_articles_from_cluster(cluster, cortar_no):
    """클러스터 내 매물 목록 조회 (페이징)"""
    all_articles = []
    page = 1
    offset = 0.06
    lat, lon = cluster["lat"], cluster["lon"]

    while True:
        data = api_get(
            "https://m.land.naver.com/cluster/ajax/articleList",
            params={
                "rletTpCd": BUILDING_TYPES, "tradTpCd": "A1", "z": "13",
                "lat": str(lat), "lon": str(lon),
                "btm": str(lat - offset), "lft": str(lon - offset),
                "top": str(lat + offset), "rgt": str(lon + offset),
                "lgeo": cluster["lgeo"], "showR0": "N",
                "page": str(page), "cortarNo": cortar_no,
            },
        )
        if data is None:
            return None
        body = data.get("body", [])
        if not body:
            break
        all_articles.extend(body)
        if not data.get("more", False):
            break
        page += 1
        wait()

    return all_articles


def fetch_all_articles_mobile(lat, lon, cortar_no, log=print):
    """한 지역의 전체 매물 수집 (클러스터 방식). 429 시 None 반환."""
    clusters = fetch_clusters(lat, lon, cortar_no)
    wait()
    if clusters is None:
        return None

    all_articles = []
    seen = set()
    for cluster in clusters:
        articles = fetch_articles_from_cluster(cluster, cortar_no)
        if articles is None:
            return None
        for a in articles:
            atcl_no = a.get("atclNo")
            if atcl_no and atcl_no not in seen:
                seen.add(atcl_no)
                all_articles.append(a)
        wait()

    return all_articles


def is_building_type(article):
    """꼬마빌딩에 해당하는 매물인지 판별"""
    tp = article.get("rletTpNm", "")
    return tp in BUILDING_TYPE_NAMES


def is_recent(article, days=1):
    """최근 N일 이내 등록된 매물인지 판별"""
    cfm = article.get("cfmYmd", "") or article.get("confirmYmd", "")
    if not cfm:
        return True  # 날짜 없으면 포함

    try:
        # PC API 날짜 형식: "20260331" 또는 "2026/03/31" 등
        clean = cfm.replace("/", "").replace(".", "").replace("-", "")
        if len(clean) == 8:
            cfm_date = datetime.strptime(clean, "%Y%m%d").date()
        elif len(clean) == 6:
            cfm_date = datetime.strptime(clean, "%y%m%d").date()
        else:
            return True
        cutoff = datetime.now().date() - timedelta(days=days)
        return cfm_date >= cutoff
    except (ValueError, TypeError):
        return True


def parse_article_mobile(article):
    """모바일 API 매물 데이터를 정리"""
    try:
        prc = int(article.get("prc", 0) or 0)
    except (ValueError, TypeError):
        prc = 0

    try:
        spc1 = float(article.get("spc1", 0) or 0)
    except (ValueError, TypeError):
        spc1 = 0
    try:
        spc2 = float(article.get("spc2", 0) or 0)
    except (ValueError, TypeError):
        spc2 = 0

    area = spc1 if spc1 > 0 else spc2
    pyeong = round(area / 3.3058, 1) if area > 0 else 0
    land_area = spc2 if spc1 > 0 and spc2 > 0 else 0
    land_pyeong = round(land_area / 3.3058, 1) if land_area > 0 else 0

    atcl_no = article.get("atclNo", "")
    link = f"https://fin.land.naver.com/articles/{atcl_no}" if atcl_no else ""

    평단가_만원 = round(prc / pyeong) if pyeong > 0 and prc > 0 else 0

    return {
        "매물번호": atcl_no,
        "매물명": article.get("atclNm", "") or article.get("bildNm", ""),
        "유형": article.get("rletTpNm", ""),
        "매매가_만원": prc,
        "매매가_억": round(prc / 10000, 1) if prc >= 10000 else f"{prc}만",
        "건물면적_m2": round(area, 1),
        "건물면적_평": pyeong,
        "토지면적_m2": round(land_area, 1),
        "토지면적_평": land_pyeong,
        "평단가_만원": 평단가_만원,
        "층": article.get("flrInfo", ""),
        "방향": article.get("tagList", [""])[0] if article.get("tagList") else "",
        "설명": article.get("atclFetrDesc", "") or "",
        "중개사": article.get("rltrNm", ""),
        "확인일": article.get("cfmYmd", "") or "",
        "링크": link,
    }


def get_sigungu_list(sido_code, log=print):
    """시도 코드 → 시군구 목록 조회 (m.land.naver.com API)"""
    try:
        resp = SESSION.get(
            "https://m.land.naver.com/map/getRegionList",
            params={"cortarNo": sido_code},
            timeout=10,
        )
        resp.raise_for_status()
        data = resp.json()
        region_list = data.get("result", {}).get("list", [])
        return [(r["CortarNm"], r["CortarNo"], float(r["MapYCrdn"]), float(r["MapXCrdn"])) for r in region_list]
    except Exception as e:
        log(f"  시군구 목록 조회 실패: {e}")
        return []


def collect_all(sido_filter=None, log=print):
    """서울 전체 수집 (시군구 단위로 조회하여 누락 방지)"""
    if sido_filter:
        codes = {k: v for k, v in TARGET_SIDO.items() if k in sido_filter}
    else:
        codes = TARGET_SIDO

    all_buildings = []
    all_skipped = []

    for sido_name, sido_code in codes.items():
        log(f"\n[{sido_name}] 시군구 목록 조회 중...")
        sigungu_list = get_sigungu_list(sido_code, log)
        wait()

        if not sigungu_list:
            log(f"  시군구 목록을 가져오지 못해 건너뜁니다.")
            continue

        sido_count = 0
        skipped = []
        for sg_name, sg_code, sg_lat, sg_lon in sigungu_list:
            raw = fetch_all_articles_mobile(sg_lat, sg_lon, sg_code, log)
            wait()

            if raw is None:
                skipped.append(sg_name)
                log(f"  {sg_name}: 429 - 180초 대기 후 재시도...")
                time.sleep(180)
                raw = fetch_all_articles_mobile(sg_lat, sg_lon, sg_code, log)
                if raw is None:
                    log(f"  {sg_name}: 재시도 실패, 스킵")
                    time.sleep(180)
                    continue

            count = 0
            for a in raw:
                if not is_building_type(a):
                    continue
                if not is_recent(a, days=RECENT_DAYS):
                    continue

                parsed = parse_article_mobile(a)
                if parsed["매매가_만원"] <= 0 or parsed["매매가_만원"] > MAX_PRICE:
                    continue

                parsed["지역_시도"] = sido_name
                parsed["지역_시군구"] = sg_name
                all_buildings.append(parsed)
                count += 1

            if count > 0:
                log(f"  {sg_name}: {count}건")
            sido_count += count

        log(f"  → {sido_name} 합계: {sido_count}건")
        if skipped:
            log(f"  [!] 스킵된 지역: {', '.join(skipped)}")
            all_skipped.extend(skipped)

    log(f"\n총 {len(all_buildings)}건 수집 완료")
    return all_buildings, all_skipped


# ── 신규 매물 판별 ────────────────────────────────────

def load_previous_ids():
    id_file = os.path.join(DATA_DIR, "collected_ids.json")
    if os.path.exists(id_file):
        with open(id_file, "r", encoding="utf-8") as f:
            return set(json.load(f))
    return set()


def save_current_ids(buildings):
    os.makedirs(DATA_DIR, exist_ok=True)
    id_file = os.path.join(DATA_DIR, "collected_ids.json")
    existing = load_previous_ids()
    current = {b["매물번호"] for b in buildings}
    merged = existing | current
    with open(id_file, "w", encoding="utf-8") as f:
        json.dump(list(merged), f, ensure_ascii=False)
    return current - existing


def is_first_run():
    """collected_ids.json이 없거나 비어있으면 첫 실행"""
    id_file = os.path.join(DATA_DIR, "collected_ids.json")
    if not os.path.exists(id_file):
        return True
    with open(id_file, "r", encoding="utf-8") as f:
        ids = json.load(f)
    return len(ids) == 0


def filter_new_buildings(buildings):
    prev_ids = load_previous_ids()
    if not prev_ids:
        return []  # 첫 실행: 기준선 수집이므로 신규 없음
    return [b for b in buildings if b["매물번호"] not in prev_ids]


# ── 엑셀 리포트 ─────────────────────────────────────

def style_header(ws, headers, row=1):
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
    border = Border(*(Side(style="thin") for _ in range(4)))
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            if cell.value:
                length = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                max_len = max(max_len, length)
        ws.column_dimensions[letter].width = min(max_len + 4, 50)


def create_report(buildings, filename, title="전체"):
    wb = Workbook()
    link_font = Font(color="0563C1", underline="single")

    ws = wb.active
    ws.title = "매물목록"

    col_headers = [
        "지역", "매물명", "유형", "매매가(억)",
        "건물면적(m2)", "건물면적(평)", "토지면적(m2)", "토지면적(평)",
        "평단가(만원/평)",
        "층", "설명", "확인일", "중개사", "네이버부동산", "링크(복사용)",
    ]
    style_header(ws, col_headers)

    sorted_buildings = sorted(buildings, key=lambda x: x["매매가_만원"])

    for i, b in enumerate(sorted_buildings, 2):
        ws.cell(row=i, column=1, value=f"{b['지역_시도']} {b['지역_시군구']}")
        ws.cell(row=i, column=2, value=b["매물명"])
        ws.cell(row=i, column=3, value=b["유형"])
        ws.cell(row=i, column=4, value=b["매매가_억"])
        ws.cell(row=i, column=5, value=b["건물면적_m2"]).number_format = "#,##0.0"
        ws.cell(row=i, column=6, value=b["건물면적_평"]).number_format = "#,##0.0"
        ws.cell(row=i, column=7, value=b["토지면적_m2"]).number_format = "#,##0.0"
        ws.cell(row=i, column=8, value=b["토지면적_평"]).number_format = "#,##0.0"
        ws.cell(row=i, column=9, value=b["평단가_만원"]).number_format = "#,##0"
        ws.cell(row=i, column=10, value=b["층"])
        ws.cell(row=i, column=11, value=b["설명"])
        ws.cell(row=i, column=12, value=b["확인일"])
        ws.cell(row=i, column=13, value=b["중개사"])

        url = b.get("링크", "")
        if url:
            cell = ws.cell(row=i, column=14, value="보기")
            cell.hyperlink = url
            cell.font = link_font
            ws.cell(row=i, column=15, value=url)

    auto_width(ws)

    # 요약 시트
    ws_sum = wb.create_sheet("요약")
    style_header(ws_sum, ["항목", "값"])

    summary = [
        ("조사 일시", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("검색 지역", title),
        ("총 매물 수", f"{len(buildings)}건"),
        ("가격 기준", "10억 이하"),
    ]

    region_counts = {}
    for b in buildings:
        key = f"{b['지역_시도']} {b['지역_시군구']}"
        region_counts[key] = region_counts.get(key, 0) + 1
    for region, cnt in sorted(region_counts.items(), key=lambda x: -x[1]):
        summary.append((region, f"{cnt}건"))

    prices = [b["매매가_만원"] for b in buildings if b["매매가_만원"] > 0]
    if prices:
        summary.append(("평균 매매가", f"{round(sum(prices)/len(prices)/10000, 1)}억원"))
        summary.append(("최저가", f"{round(min(prices)/10000, 1)}억원"))
        summary.append(("최고가", f"{round(max(prices)/10000, 1)}억원"))

    단가들 = [b["평단가_만원"] for b in buildings if b["평단가_만원"] > 0]
    if 단가들:
        summary.append(("평균 평단가", f"{round(sum(단가들)/len(단가들)):,}만원/평"))

    for i, (label, value) in enumerate(summary, 2):
        ws_sum.cell(row=i, column=1, value=label)
        ws_sum.cell(row=i, column=2, value=value)
    auto_width(ws_sum)

    wb.save(filename)
    return filename


# ── 텔레그램 알림 ────────────────────────────────────

def send_telegram(text):
    try:
        requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
            json={
                "chat_id": TELEGRAM_CHAT_ID,
                "text": text,
                "parse_mode": "HTML",
                "disable_web_page_preview": True,
            },
            timeout=10,
        )
    except Exception:
        pass


def send_telegram_file(filepath, caption=""):
    """텔레그램으로 파일 전송"""
    try:
        with open(filepath, "rb") as f:
            requests.post(
                f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendDocument",
                data={"chat_id": TELEGRAM_CHAT_ID, "caption": caption},
                files={"document": f},
                timeout=30,
            )
    except Exception:
        pass


def notify_buildings(buildings, excel_path=None):
    if not buildings:
        send_telegram("오늘 신규 꼬마빌딩 매물이 없습니다.")
        return

    if excel_path and os.path.exists(excel_path):
        send_telegram_file(excel_path)
    else:
        send_telegram(f"꼬마빌딩 {len(buildings)}건 수집 완료")


# ── 메인 ─────────────────────────────────────────────

def main():
    global RECENT_DAYS

    region_filter = None
    if "--region" in sys.argv:
        idx = sys.argv.index("--region")
        if idx + 1 < len(sys.argv):
            region_filter = [sys.argv[idx + 1]]

    if "--days" in sys.argv:
        idx = sys.argv.index("--days")
        if idx + 1 < len(sys.argv):
            RECENT_DAYS = int(sys.argv[idx + 1])

    notify = "--notify" in sys.argv

    print("=" * 50)
    print("  소액 꼬마빌딩 급매물 수집기 v2")
    print("  10억 이하 | 서울")
    print("=" * 50)
    print()

    try:
        buildings, skipped = collect_all(sido_filter=region_filter)
    except Exception as e:
        print(f"\n수집 중 오류: {e}")
        if notify:
            send_telegram(f"[오류] 급매물 수집 실패\n{e}")
        return

    if not buildings:
        msg = "오늘 신규 꼬마빌딩 매물이 없습니다."
        if skipped:
            msg += f"\n\n[!] 429 차단으로 {len(skipped)}개 지역 수집 실패:\n{', '.join(skipped)}"
        print(f"\n{msg}")
        if notify:
            send_telegram(msg)
        return

    first_run = is_first_run()
    new_buildings = filter_new_buildings(buildings)
    save_current_ids(buildings)

    if first_run:
        print(f"\n[기준선 수집] 전체 {len(buildings)}건 등록 완료")
        print("  다음 실행부터 신규 매물만 따로 분류됩니다.")
    else:
        print(f"\n신규 매물: {len(new_buildings)}건 / 전체: {len(buildings)}건")

    os.makedirs(DB_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    region_label = region_filter[0] if region_filter else "서울"

    filename = os.path.join(DB_DIR, f"급매물_{region_label}_{timestamp}.xlsx")
    create_report(buildings, filename, title=region_label)
    print(f"\n전체 리포트 저장: {filename}")

    if not first_run and new_buildings:
        new_filename = os.path.join(DB_DIR, f"신규매물_{region_label}_{timestamp}.xlsx")
        create_report(new_buildings, new_filename, title=f"{region_label} 신규")
        print(f"신규 리포트 저장: {new_filename}")

    if notify:
        if first_run:
            msg = f"[기준선 수집] 꼬마빌딩 {len(buildings)}건 등록 완료\n다음부터 신규 매물 알림이 시작됩니다."
            if skipped:
                msg += f"\n\n[!] {len(skipped)}개 지역 429 스킵: {', '.join(skipped)}"
            send_telegram(msg)
        else:
            notify_buildings(new_buildings, excel_path=filename)
            if skipped:
                send_telegram(f"[!] {len(skipped)}개 지역 429 스킵: {', '.join(skipped)}")
        print("텔레그램 알림 발송 완료")

    print("\n완료!")


if __name__ == "__main__":
    main()
